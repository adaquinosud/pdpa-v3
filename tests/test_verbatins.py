"""Testes da API + UI de verbatins (Bloco 4 CP-C)."""

from __future__ import annotations

from datetime import datetime, timedelta

from src.models.verbatim import Verbatim
from src.models.verbatim_reclassificacao import VerbatimReclassificacao


def _criar_verbatim(
    db_session,
    empresa_id,
    fonte_id,
    local_id=None,
    texto="texto exemplo",
    subpilar="Pa1",
    tipo="promotor",
    confianca=0.85,
    justificativa=None,
    data_dias_atras=10,
    autor=None,
):
    v = Verbatim(
        empresa_id=empresa_id,
        fonte_id=fonte_id,
        local_id=local_id,
        texto=texto,
        autor=autor,
        data_criacao_original=datetime.utcnow() - timedelta(days=data_dias_atras),
        hash_dedup=f"hash-{texto[:30]}",
        subpilar=subpilar,
        tipo=tipo,
        confianca=confianca,
        justificativa=justificativa,
    )
    db_session.add(v)
    db_session.commit()
    return v


def _empresa_com_estrutura(client_loyall, suffix=None):
    """Cria empresa + agrupamento + local + fonte; devolve dict com ids.

    Suffix opcional pra evitar colisão de nomes em testes que chamam
    a função várias vezes (UNIQUE empresas.nome).
    """
    import uuid

    sfx = suffix or uuid.uuid4().hex[:6]
    e = client_loyall.post("/api/empresas/", json={"nome": f"EVrb-{sfx}"}).get_json()
    a = client_loyall.post(
        f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "GVrb"}
    ).get_json()
    loc = client_loyall.post(
        f"/api/empresas/{e['id']}/locais",
        json={"nome": "LVrb", "agrupamento_id": a["id"]},
    ).get_json()
    f = client_loyall.post(
        f"/api/locais/{loc['id']}/fontes",
        json={"conector_tipo": "google", "url": f"ChIJ_v_{sfx}"},
    ).get_json()
    return {"e": e, "a": a, "loc": loc, "f": f}


# ── API: lista paginada + filtros ────────────────────────────────────────


def test_api_listar_verbatins_vazio(client_loyall):
    e = client_loyall.post("/api/empresas/", json={"nome": "EvVazio"}).get_json()
    r = client_loyall.get(f"/api/empresas/{e['id']}/verbatins")
    assert r.status_code == 200
    body = r.get_json()
    assert body["total"] == 0
    assert body["verbatins"] == []
    assert body["pagina"] == 1
    assert body["por_pagina"] == 50


def test_api_listar_verbatins_com_dados(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        empresa_id=ctx["e"]["id"],
        fonte_id=ctx["f"]["id"],
        local_id=ctx["loc"]["id"],
        texto="Atendimento excelente",
        subpilar="Pa1",
        tipo="promotor",
        justificativa="Elogio direto ao atendimento.",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins")
    body = r.get_json()
    assert body["total"] == 1
    v = body["verbatins"][0]
    assert v["texto"] == "Atendimento excelente"
    assert v["subpilar"] == "Pa1"
    assert v["agrupamento_nome"] == "GVrb"
    assert v["local_nome"] == "LVrb"
    assert v["fonte_conector_tipo"] == "google"
    assert v["justificativa"] == "Elogio direto ao atendimento."


def test_api_filtro_subpilar(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t1",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t2",
        subpilar="D2",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?subpilar=D2")
    assert r.get_json()["total"] == 1
    assert r.get_json()["verbatins"][0]["subpilar"] == "D2"


def test_api_filtro_tipo(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="elogio",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="critica",
        subpilar="D2",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?tipo=detrator")
    assert r.get_json()["total"] == 1


def test_api_filtro_busca_texto(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="Camarão delicioso",
        subpilar="P2",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="Atendimento ruim",
        subpilar="Pa1",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?q=camar%C3%A3o")
    assert r.get_json()["total"] == 1
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?q=atendimento")
    assert r.get_json()["total"] == 1


def test_api_filtro_agrupamento_e_local(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    # Cria outro agrupamento + local da mesma empresa
    a2 = client_loyall.post(
        f"/api/empresas/{ctx['e']['id']}/agrupamentos", json={"nome": "G2"}
    ).get_json()
    loc2 = client_loyall.post(
        f"/api/empresas/{ctx['e']['id']}/locais",
        json={"nome": "L2", "agrupamento_id": a2["id"]},
    ).get_json()
    f2 = client_loyall.post(
        f"/api/locais/{loc2['id']}/fontes",
        json={"conector_tipo": "google", "url": "ChIJ_2"},
    ).get_json()
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="v1",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        f2["id"],
        loc2["id"],
        texto="v2",
    )

    # Filtra por agrupamento G2 → só v2
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?agrupamento_id={a2['id']}")
    assert r.get_json()["total"] == 1
    assert r.get_json()["verbatins"][0]["texto"] == "v2"

    # Filtra por local LVrb → só v1
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?local_id={ctx['loc']['id']}")
    assert r.get_json()["total"] == 1
    assert r.get_json()["verbatins"][0]["texto"] == "v1"


def test_api_paginacao(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    for i in range(7):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"v{i}",
            data_dias_atras=20 - i,
        )
    # 7 itens, 3 por página
    r1 = client_loyall.get(
        f"/api/empresas/{ctx['e']['id']}/verbatins?por_pagina=3&pagina=1"
    ).get_json()
    assert r1["total"] == 7
    assert len(r1["verbatins"]) == 3
    r3 = client_loyall.get(
        f"/api/empresas/{ctx['e']['id']}/verbatins?por_pagina=3&pagina=3"
    ).get_json()
    assert len(r3["verbatins"]) == 1


# ── PATCH /reclassificar ────────────────────────────────────────────────


def test_api_reclassificar_atualiza_e_cria_historico(client_loyall, db_session, usuario_loyall):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        subpilar="Pa1",
        tipo="promotor",
    )
    r = client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "D2", "tipo": "detrator", "justificativa": "rever"},
    )
    assert r.status_code == 200
    body = r.get_json()
    assert body["subpilar"] == "D2"
    assert body["tipo"] == "detrator"
    assert body["subpilar_anterior"] == "Pa1"
    assert body["tipo_anterior"] == "promotor"

    # Histórico criado
    db_session.expire_all()
    histo = db_session.query(VerbatimReclassificacao).filter_by(verbatim_id=v.id).all()
    assert len(histo) == 1
    assert histo[0].subpilar_anterior == "Pa1"
    assert histo[0].subpilar_novo == "D2"
    assert histo[0].justificativa == "rever"


def test_api_reclassificar_valida_subpilar(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"])
    r = client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "XX", "tipo": "promotor"},
    )
    assert r.status_code == 400


def test_api_reclassificar_valida_sem_lastro_inativo(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"])
    # sem_lastro com promotor → 400
    r = client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "sem_lastro", "tipo": "promotor"},
    )
    assert r.status_code == 400


def test_api_reclassificar_multiplas_vezes_acumula_historico(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        subpilar="Pa1",
        tipo="promotor",
    )
    client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "D2", "tipo": "detrator"},
    )
    client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "P2", "tipo": "conversivel"},
    )
    db_session.expire_all()
    histo = (
        db_session.query(VerbatimReclassificacao)
        .filter_by(verbatim_id=v.id)
        .order_by(VerbatimReclassificacao.id)
        .all()
    )
    assert len(histo) == 2
    assert histo[0].subpilar_novo == "D2"
    assert histo[1].subpilar_anterior == "D2"
    assert histo[1].subpilar_novo == "P2"


# ── DELETE ──────────────────────────────────────────────────────────────


def test_api_delete_verbatim(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"])
    r = client_loyall.delete(f"/api/verbatins/{v.id}")
    assert r.status_code == 200
    assert client_loyall.get(f"/api/verbatins/{v.id}").status_code == 404


# ── Isolamento por papel ────────────────────────────────────────────────


def test_cliente_so_ve_verbatins_da_sua_empresa(client_loyall, client_cliente_factory, db_session):
    ctx_a = _empresa_com_estrutura(client_loyall)
    ctx_b = _empresa_com_estrutura(client_loyall)
    # Renomeia para evitar colisão de nomes
    client_loyall.put(f"/api/empresas/{ctx_b['e']['id']}", json={"nome": "EVrbOutra"})
    _criar_verbatim(
        db_session,
        ctx_a["e"]["id"],
        ctx_a["f"]["id"],
        ctx_a["loc"]["id"],
        texto="da empresa A",
    )
    _criar_verbatim(
        db_session,
        ctx_b["e"]["id"],
        ctx_b["f"]["id"],
        ctx_b["loc"]["id"],
        texto="da empresa B",
    )
    cli = client_cliente_factory(ctx_a["e"]["id"])
    # Cliente acessa só os da própria empresa
    r = cli.get(f"/api/empresas/{ctx_a['e']['id']}/verbatins")
    assert r.status_code == 200
    assert r.get_json()["total"] == 1
    # E NÃO acessa os da outra
    r2 = cli.get(f"/api/empresas/{ctx_b['e']['id']}/verbatins")
    assert r2.status_code == 403


def test_cliente_pode_reclassificar_da_propria_empresa(
    client_loyall, client_cliente_factory, db_session
):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        subpilar="Pa1",
        tipo="promotor",
    )
    cli = client_cliente_factory(ctx["e"]["id"])
    r = cli.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "D2", "tipo": "detrator"},
    )
    assert r.status_code == 200


def test_cliente_nao_reclassifica_de_outra_empresa(
    client_loyall, client_cliente_factory, db_session
):
    ctx_a = _empresa_com_estrutura(client_loyall)
    ctx_b = _empresa_com_estrutura(client_loyall)
    client_loyall.put(f"/api/empresas/{ctx_b['e']['id']}", json={"nome": "EVrbOutra"})
    v_b = _criar_verbatim(
        db_session,
        ctx_b["e"]["id"],
        ctx_b["f"]["id"],
        ctx_b["loc"]["id"],
    )
    cli = client_cliente_factory(ctx_a["e"]["id"])
    r = cli.patch(
        f"/api/verbatins/{v_b.id}/reclassificar",
        json={"subpilar": "Pa1", "tipo": "promotor"},
    )
    assert r.status_code == 403


# ── UI ──────────────────────────────────────────────────────────────────


def test_ui_pagina_verbatins_renderiza(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="texto na lista",
        subpilar="Pa1",
        tipo="promotor",
    )
    r = client_loyall.get(f"/empresas/{ctx['e']['id']}/verbatins")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "texto na lista" in html
    assert "Pa1" in html
    assert 'name="q"' in html  # filtro busca
    assert 'name="subpilar"' in html


def test_ui_modal_reclassificar(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"])
    r = client_loyall.get(f"/ui/verbatins/{v.id}/reclassificar")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Reclassificar" in html
    assert 'name="subpilar"' in html
    assert 'name="tipo"' in html
    assert 'name="justificativa"' in html


def test_ui_salvar_reclassificacao(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        subpilar="Pa1",
        tipo="promotor",
    )
    r = client_loyall.patch(
        f"/ui/verbatins/{v.id}/reclassificar",
        data={"subpilar": "D2", "tipo": "detrator", "justificativa": "via ui"},
    )
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "D2" in html
    assert "detrator" in html
    # Verifica persistência
    body = client_loyall.get(f"/api/verbatins/{v.id}").get_json()
    assert body["subpilar"] == "D2"


def test_ui_modal_detalhes_mostra_historico(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        subpilar="Pa1",
        tipo="promotor",
    )
    client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "D2", "tipo": "detrator", "justificativa": "1ª revisão"},
    )
    r = client_loyall.get(f"/ui/verbatins/{v.id}/detalhes")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Histórico" in html
    assert "1ª revisão" in html


def test_ui_excluir_verbatim(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"])
    r = client_loyall.delete(f"/ui/verbatins/{v.id}")
    assert r.status_code == 200
    assert client_loyall.get(f"/api/verbatins/{v.id}").status_code == 404


def test_ui_pagina_verbatins_cliente_da_outra_empresa_403(client_loyall, client_cliente_factory):
    e_a = client_loyall.post("/api/empresas/", json={"nome": "EVrbA"}).get_json()
    e_b = client_loyall.post("/api/empresas/", json={"nome": "EVrbB"}).get_json()
    cli = client_cliente_factory(e_a["id"])
    r = cli.get(f"/empresas/{e_b['id']}/verbatins")
    assert r.status_code == 403


# ── Exportar XLSX ────────────────────────────────────────────────────────


def test_exportar_xlsx_basico(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="Verbatim 1",
        subpilar="Pa1",
        tipo="promotor",
        justificativa="Boa",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="Verbatim 2",
        subpilar="D2",
        tipo="detrator",
        justificativa="Ruim",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins/exportar.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    assert "attachment" in r.headers.get("Content-Disposition", "").lower()
    assert r.data[:2] == b"PK"  # XLSX é zip

    # Confere conteúdo do XLSX
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "id"
    assert "texto" in rows[0]
    assert "review_id_externo" in rows[0]
    assert "historico_reclassificacoes" in rows[0]
    assert len(rows) == 3  # 1 header + 2 verbatins
    textos = [r[5] for r in rows[1:]]
    assert "Verbatim 1" in textos and "Verbatim 2" in textos


def test_exportar_xlsx_aplica_filtros(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="só esse",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="outro",
        subpilar="D2",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins/exportar.xlsx?subpilar=Pa1")
    assert r.status_code == 200
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # header + 1 verbatim filtrado
    assert rows[1][5] == "só esse"
    assert rows[1][6] == "Pa1"


def test_exportar_xlsx_cliente_outra_empresa_403(client_loyall, client_cliente_factory):
    e_a = client_loyall.post("/api/empresas/", json={"nome": "EVrbXlsxA"}).get_json()
    e_b = client_loyall.post("/api/empresas/", json={"nome": "EVrbXlsxB"}).get_json()
    cli = client_cliente_factory(e_a["id"])
    r = cli.get(f"/api/empresas/{e_b['id']}/verbatins/exportar.xlsx")
    assert r.status_code == 403


def test_exportar_xlsx_excede_cap_devolve_413(client_loyall, db_session, monkeypatch):
    ctx = _empresa_com_estrutura(client_loyall)
    _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"], texto="t")
    monkeypatch.setattr("src.api.verbatins.EXPORTAR_XLSX_MAX_LINHAS", 0)
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins/exportar.xlsx")
    assert r.status_code == 413
    body = r.get_json()
    assert "limite" in body
    assert body["limite"] == 0
    assert body["total"] == 1


def test_filtro_sem_classificacao_pega_verbatins_subpilar_null(client_loyall, db_session):
    """B5 ext. CP-2: ?sem_classificacao=1 filtra verbatins com subpilar IS NULL.

    Necessário para link "sem classificação" no painel funcionar.
    """
    ctx = _empresa_com_estrutura(client_loyall)
    # 1 verbatim classificado, 1 sem classificação (subpilar=NULL)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="classificado",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="sem_classif",
        subpilar=None,
        tipo=None,
    )
    # Sem filtro: 2
    r_all = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins")
    assert r_all.get_json()["total"] == 2
    # Com filtro sem_classificacao=1: 1
    r_filt = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins?sem_classificacao=1")
    body = r_filt.get_json()
    assert body["total"] == 1
    assert body["verbatins"][0]["texto"] == "sem_classif"
    assert body["verbatins"][0]["subpilar"] is None


def test_exportar_xlsx_inclui_historico_reclassificacoes(client_loyall, db_session):
    ctx = _empresa_com_estrutura(client_loyall)
    v = _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="reclassificado",
        subpilar="Pa1",
        tipo="promotor",
    )
    client_loyall.patch(
        f"/api/verbatins/{v.id}/reclassificar",
        json={"subpilar": "D1", "tipo": "detrator", "motivo": "ajuste"},
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/verbatins/exportar.xlsx")
    assert r.status_code == 200
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hist = rows[1][-1]  # última coluna
    assert "Pa1→D1" in hist
    assert "promotor→detrator" in hist
