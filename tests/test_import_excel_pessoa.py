"""Tests do import Excel no modo 'interno identificado' (cria Pessoa).

Frente aditiva sobre fluxo vivo: modo DESLIGADO = import idêntico ao de hoje
(sem Pessoa). Modo LIGADO = cria Pessoa(interno_consentido) por email|id_cliente,
fonte 'excel_interno', com gate de consentimento.
"""

from __future__ import annotations

import json

import pandas as pd

from src.coletor.excel import _detectar_colunas, gerar_modelo_xlsx, importar_arquivo
from src.models.empresa import Empresa
from src.models.fonte import Fonte
from src.models.pessoa import Pessoa, PessoaIdentificador
from src.models.verbatim import Verbatim


def _empresa(db_session, nome):
    e = Empresa(nome=nome)
    db_session.add(e)
    db_session.commit()
    return e.id


def _csv(tmp_path, rows, nome="t.csv"):
    arq = tmp_path / nome
    pd.DataFrame(rows).to_csv(arq, index=False)
    return arq


# ── Detecção de colunas ──────────────────────────────────────────────────────


def test_deteccao_normal_nao_ve_identidade():
    """Modo normal (byte-a-byte de hoje): sem campos de identidade no mapa."""
    m = _detectar_colunas(["texto", "email", "id_cliente"])
    assert "email" not in m and "id_cliente" not in m
    assert m["texto"] == "texto"


def test_deteccao_interna_pega_email_e_id():
    m = _detectar_colunas(["Comentário", "E-mail", "Codigo Cliente"], interno=True)
    assert m["texto"] == "Comentário"
    assert m["email"] == "E-mail"
    assert m["id_cliente"] == "Codigo Cliente"


# ── Modo desligado = regressão zero ──────────────────────────────────────────


def test_modo_desligado_identico(db_session, tmp_path):
    e = _empresa(db_session, "EDesligado")
    arq = _csv(tmp_path, [{"texto": "Bom atendimento", "email": "a@x.com", "autor": "Ana"}])
    stats = importar_arquivo(arq, empresa_id=e)  # interno desligado (default)
    assert stats["importados"] == 1
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    assert v.pessoa_id is None and v.autor == "Ana"  # sem Pessoa; autor intacto
    assert db_session.query(Pessoa).count() == 0
    f = db_session.get(Fonte, v.fonte_id)
    assert f.conector_tipo == "excel_manual"  # regime normal


# ── Gate de consentimento ────────────────────────────────────────────────────


def test_interno_sem_consentimento_bloqueia(db_session, tmp_path):
    e = _empresa(db_session, "ESemConsent")
    arq = _csv(tmp_path, [{"texto": "x", "email": "a@x.com"}])
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=False)
    assert stats["importados"] == 0 and stats["erros_validacao"]
    assert db_session.query(Verbatim).filter_by(empresa_id=e).count() == 0


def test_interno_sem_coluna_identidade_valida(db_session, tmp_path):
    e = _empresa(db_session, "ESemId")
    arq = _csv(tmp_path, [{"texto": "x", "autor": "Ana"}])  # sem email nem id_cliente
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["importados"] == 0
    assert any("email ou id_cliente" in m for m in stats["erros_validacao"])


# ── Modo ligado: cria Pessoa ─────────────────────────────────────────────────


def test_interno_cria_pessoa_e_compartilha(db_session, tmp_path):
    e = _empresa(db_session, "EInterno")
    arq = _csv(
        tmp_path,
        [
            {"texto": "Adorei", "email": "Joao@X.com", "autor": "João"},
            {"texto": "Voltarei", "email": "joao@x.com", "autor": "João"},  # mesma pessoa
            {"texto": "Outra", "email": "maria@x.com", "autor": "Maria"},
        ],
    )
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["importados"] == 3
    # 2 pessoas (joão dedup por email normalizado, maria)
    assert db_session.query(Pessoa).count() == 2
    vs = db_session.query(Verbatim).filter_by(empresa_id=e).all()
    por_email = {v.texto: v.pessoa_id for v in vs}
    assert por_email["Adorei"] == por_email["Voltarei"]  # mesma Pessoa
    assert por_email["Outra"] != por_email["Adorei"]
    # identificador correto
    ident = db_session.query(PessoaIdentificador).filter_by(external_id="joao@x.com").one()
    # email agora vai pra fonte 'pesquisa' (mesmo namespace da pesquisa-PDPA) — não 'excel'
    assert ident.tipo == "interno_consentido" and ident.fonte == "pesquisa"
    assert json.loads(ident.atributos_json)["opt_in"] is True
    # fonte interna
    f = db_session.get(Fonte, vs[0].fonte_id)
    assert f.conector_tipo == "excel_interno" and f.autenticacao_tipo == "autenticada"


def test_interno_fallback_id_cliente(db_session, tmp_path):
    e = _empresa(db_session, "EFallback")
    arq = _csv(tmp_path, [{"texto": "Oi", "id_cliente": "CRM-42", "autor": "Zé"}])
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["importados"] == 1
    ident = db_session.query(PessoaIdentificador).one()
    assert ident.external_id == "CRM-42"


def test_interno_linha_sem_identidade_nao_quebra(db_session, tmp_path):
    e = _empresa(db_session, "EMisto")
    arq = _csv(
        tmp_path,
        [
            {"texto": "Com email", "email": "a@x.com"},
            {"texto": "Sem nada", "email": ""},  # sem identidade
        ],
    )
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["importados"] == 2 and stats["sem_identidade"] == 1
    vs = {v.texto: v.pessoa_id for v in db_session.query(Verbatim).filter_by(empresa_id=e)}
    assert vs["Com email"] is not None and vs["Sem nada"] is None


def test_interno_reimport_nao_duplica_pessoa(db_session, tmp_path):
    e = _empresa(db_session, "EReimport")
    rows = [{"texto": "Único", "email": "a@x.com", "autor": "A"}]
    arq = _csv(tmp_path, rows)
    importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)  # re-import
    assert db_session.query(Pessoa).count() == 1  # UNIQUE garante idempotência
    assert db_session.query(PessoaIdentificador).filter_by(external_id="a@x.com").count() == 1


# ── Frente 2: identidade unificada (import cruza com a pesquisa-PDPA) ─────────


def test_import_colapsa_com_pessoa_da_pesquisa(db_session, tmp_path):
    """O ponto da Frente 2: o mesmo id_cliente do CSAT importado e de uma pesquisa-PDPA
    viram a MESMA Pessoa (fonte 'crm' compartilhada) — cruza fontes por pessoa."""
    from src.coletor.excel import _reconciliar_pessoa

    e = _empresa(db_session, "ECross")
    # simula uma resposta de pesquisa que já criou a Pessoa por id_cliente (CRM é chave
    # POR EMPRESA — §5.5 —, então o seed carrega a MESMA empresa do import).
    pid_pesq = _reconciliar_pessoa(
        db_session, id_cliente="CRM-99", nome="Cliente X", origem="pesquisa_web", empresa_id=e
    )
    db_session.commit()
    # importa um CSAT com o MESMO id_cliente
    arq = _csv(tmp_path, [{"texto": "importado", "id_cliente": "CRM-99", "autor": "X"}])
    importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    assert v.pessoa_id == pid_pesq  # colapsou na Pessoa da pesquisa
    assert (
        db_session.query(PessoaIdentificador).filter_by(fonte="crm", external_id="CRM-99").count()
        == 1
    )


def test_import_email_colapsa_com_pesquisa(db_session, tmp_path):
    """Mesmo email (fonte 'pesquisa', normalizado lower nos dois canais) colapsa também."""
    from src.coletor.excel import _reconciliar_pessoa

    e = _empresa(db_session, "ECrossMail")
    pid = _reconciliar_pessoa(db_session, email="ana@x.com", origem="pesquisa_web")
    db_session.commit()
    arq = _csv(tmp_path, [{"texto": "oi", "email": "Ana@X.com", "autor": "Ana"}])  # case diferente
    importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    assert v.pessoa_id == pid  # mesma Pessoa (email normalizado igual)


def test_import_multichave_funde_pessoas(db_session, tmp_path):
    """Linha com email+id_cliente que já apontavam pra Pessoas distintas → funde numa só;
    a stat pessoas_merges registra a fusão (auditoria)."""
    from src.coletor.excel import _reconciliar_pessoa

    e = _empresa(db_session, "EMerge")
    # e-mail é chave GLOBAL; CRM é POR EMPRESA (§5.5) — o CRM pré-existente nasce na
    # empresa do import, senão o lookup escopado não o reencontraria pra fundir.
    pa = _reconciliar_pessoa(db_session, email="dupla@x.com", origem="pesquisa_web")
    pb = _reconciliar_pessoa(db_session, id_cliente="CRM-7", origem="pesquisa_web", empresa_id=e)
    db_session.commit()
    assert pa != pb  # começam separadas
    arq = _csv(
        tmp_path,
        [{"texto": "eu", "email": "dupla@x.com", "id_cliente": "CRM-7", "autor": "D"}],
    )
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["pessoas_merges"] == 1  # fundiu 1 par pré-existente
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    id_email = (
        db_session.query(PessoaIdentificador)
        .filter_by(fonte="pesquisa", external_id="dupla@x.com")
        .one()
    )
    id_crm = db_session.query(PessoaIdentificador).filter_by(fonte="crm", external_id="CRM-7").one()
    assert id_email.pessoa_id == id_crm.pessoa_id == v.pessoa_id


def test_import_nunca_funde_por_nome(db_session, tmp_path):
    """Duas linhas com o MESMO autor mas id_cliente diferentes → 2 Pessoas distintas.
    Nome é só rótulo de exibição, NUNCA chave de fusão."""
    e = _empresa(db_session, "ENome")
    arq = _csv(
        tmp_path,
        [
            {"texto": "a", "id_cliente": "CRM-A", "autor": "Maria Souza"},
            {"texto": "b", "id_cliente": "CRM-B", "autor": "Maria Souza"},
        ],
    )
    stats = importar_arquivo(arq, empresa_id=e, interno_identificado=True, consentimento=True)
    assert stats["pessoas_merges"] == 0
    assert stats["pessoas_vinculadas"] == 2  # duas Pessoas distintas
    vs = db_session.query(Verbatim).filter_by(empresa_id=e).all()
    assert len({v.pessoa_id for v in vs}) == 2  # não fundiu por nome


# ── Modelo para download (Tarefa 5) ──────────────────────────────────────────


def test_modelo_xlsx_colunas_por_modo():
    """Normal: base + grão (local/agrupamento). Interno: + email/id_cliente. Contrato
    visual — cada cabeçalho é nome canônico que o detector casa."""
    from src.coletor.excel import _aliases_efetivos

    normal = pd.read_excel(gerar_modelo_xlsx(interno_identificado=False))
    assert list(normal.columns) == ["texto", "rating", "data", "autor", "local", "agrupamento"]
    assert len(normal) >= 1
    interno = pd.read_excel(gerar_modelo_xlsx(interno_identificado=True))
    assert "email" in interno.columns and "id_cliente" in interno.columns
    # cada cabeçalho do modelo é um nome que a detecção reconhece (canônico ∈ seu alias)
    for modo, df in ((False, normal), (True, interno)):
        aliases = _aliases_efetivos(modo)
        for col in df.columns:
            assert col in aliases[col], f"cabeçalho {col!r} não casa o detector"


def test_modelo_xlsx_import_le_primeira_aba():
    """A 2ª aba (instruções) não atrapalha: o import lê SEMPRE a 1ª aba (os dados)."""
    from src.coletor.excel import _detectar_colunas

    df = pd.read_excel(gerar_modelo_xlsx(interno_identificado=True))  # 1ª aba
    cols = _detectar_colunas(list(df.columns), interno=True)
    # detecta grão + identidade a partir dos cabeçalhos do modelo
    assert cols["local"] and cols["agrupamento"] and cols["email"] and cols["id_cliente"]


def test_rota_modelo_download(client_loyall):
    r = client_loyall.get("/importar-verbatins/modelo")
    assert r.status_code == 200
    assert "attachment" in r.headers.get("Content-Disposition", "")
    assert "modelo_import.xlsx" in r.headers.get("Content-Disposition", "")
    ri = client_loyall.get("/importar-verbatins/modelo?interno=1")
    assert ri.status_code == 200
    assert "modelo_import_interno.xlsx" in ri.headers.get("Content-Disposition", "")


def test_rota_modelo_interno_entrega_identidade(client_loyall):
    """Contrato da ROTA (não só do gerador): ?interno=1 devolve email/id_cliente; sem o
    param, não. O teste antigo só via o filename — por isso o bug do link passou."""
    import io

    normal = pd.read_excel(io.BytesIO(client_loyall.get("/importar-verbatins/modelo").data))
    assert "email" not in normal.columns and "id_cliente" not in normal.columns
    interno = pd.read_excel(
        io.BytesIO(client_loyall.get("/importar-verbatins/modelo?interno=1").data)
    )
    assert "email" in interno.columns and "id_cliente" in interno.columns


def test_rota_modelo_por_empresa_popula_dropdown(client_loyall, db_session):
    """Caminho HTTP REAL: GET com empresa_id → o xlsx traz dropdown FECHADO de local
    referenciando os locais DA EMPRESA (na aba oculta 'listas'), + trava 1-5 e data. Testa
    o request→xlsx (não o gerador isolado) — foi o furo do bug do link."""
    import io

    from openpyxl import load_workbook

    from src.models.agrupamento import Agrupamento
    from src.models.local import Local

    e = _empresa(db_session, "EDrop")
    db_session.add_all(
        [
            Local(empresa_id=e, nome="Loja Centro"),
            Local(empresa_id=e, nome="Loja Sul"),
            Agrupamento(empresa_id=e, nome="Vendas"),
        ]
    )
    db_session.commit()

    r = client_loyall.get(f"/importar-verbatins/modelo?empresa_id={e}")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data))
    assert "listas" in wb.sheetnames and wb["listas"].sheet_state == "hidden"
    assert [wb["listas"][f"A{i}"].value for i in (1, 2)] == ["Loja Centro", "Loja Sul"]
    assert wb["listas"]["B1"].value == "Vendas"
    dvs = wb["verbatins"].data_validations.dataValidation
    tipos = {dv.type for dv in dvs}
    assert {"list", "whole", "date"} <= tipos  # dropdown + trava numérica + data
    list_dv = next(dv for dv in dvs if dv.type == "list")
    assert "listas!" in list_dv.formula1  # dropdown fechado referenciando a aba oculta


def test_rota_modelo_empresa_sem_locais_sem_dropdown(client_loyall, db_session):
    """Empresa nova (sem locais) → modelo ainda gerado (rating/data travados) mas SEM
    dropdown de local — o import cai empresa-wide até cadastrar locais."""
    import io

    from openpyxl import load_workbook

    e = _empresa(db_session, "ESemLocal")
    r = client_loyall.get(f"/importar-verbatins/modelo?empresa_id={e}")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data))
    tipos = [dv.type for dv in wb["verbatins"].data_validations.dataValidation]
    assert "list" not in tipos  # sem dropdown (sem locais/agrupamentos)
    assert "whole" in tipos and "date" in tipos  # rating/data seguem travados


# ── Fatia 2: guard do backend (import não cria local/agrupamento desconhecido) ─


def test_guard_local_desconhecido_pula(db_session, tmp_path):
    """Empresa COM locais: local preenchido-e-inexistente PULA a linha + conta no aviso;
    NÃO cria o local (corrigível: cadastra + reimporta no grão certo)."""
    from src.models.local import Local

    e = _empresa(db_session, "EGuard")
    db_session.add(Local(empresa_id=e, nome="Loja Centro"))
    db_session.commit()
    arq = _csv(
        tmp_path,
        [
            {"texto": "boa", "rating": 5, "local": "Loja Centro"},  # conhecido → entra
            {"texto": "ruim", "rating": 2, "local": "Loja Fantasma"},  # desconhecido → pula
        ],
    )
    stats = importar_arquivo(arq, empresa_id=e)
    assert stats["importados"] == 1 and stats["linhas_local_desconhecido"] == 1
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    loc = db_session.query(Local).filter_by(empresa_id=e, nome="Loja Centro").one()
    assert v.local_id == loc.id  # a linha conhecida entrou no grão
    assert db_session.query(Local).filter_by(empresa_id=e).count() == 1  # não criou o fantasma


def test_guard_empresa_nova_cai_empresa_wide(db_session, tmp_path):
    """Empresa nova (0 locais): guard OFF — local preenchido cai empresa-wide (importa,
    não pula, não cria) — senão o 1º import nunca aconteceria."""
    from src.models.local import Local

    e = _empresa(db_session, "ENova")
    arq = _csv(tmp_path, [{"texto": "oi", "rating": 4, "local": "Qualquer"}])
    stats = importar_arquivo(arq, empresa_id=e)
    assert stats["importados"] == 1 and stats["linhas_local_desconhecido"] == 0
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    assert v.local_id is None  # empresa-wide
    assert db_session.query(Local).filter_by(empresa_id=e).count() == 0  # nada criado


def test_guard_local_vazio_empresa_wide(db_session, tmp_path):
    """Local VAZIO (mesmo com a empresa tendo locais) = empresa-wide legítimo — não pula."""
    from src.models.local import Local

    e = _empresa(db_session, "EVazio")
    db_session.add(Local(empresa_id=e, nome="Loja X"))
    db_session.commit()
    arq = _csv(tmp_path, [{"texto": "sem local", "rating": 3, "local": ""}])
    stats = importar_arquivo(arq, empresa_id=e)
    assert stats["importados"] == 1 and stats["linhas_local_desconhecido"] == 0
    v = db_session.query(Verbatim).filter_by(empresa_id=e).one()
    assert v.local_id is None  # empresa-wide


def test_guard_agrupamento_desconhecido_nao_pula(db_session, tmp_path):
    """Agrupamento desconhecido NÃO pula (o grão é o local) — só conta no aviso; não cria."""
    from src.models.agrupamento import Agrupamento
    from src.models.local import Local

    e = _empresa(db_session, "EAgr")
    db_session.add(Local(empresa_id=e, nome="Loja Y"))
    db_session.commit()
    arq = _csv(
        tmp_path,
        [{"texto": "ok", "rating": 5, "local": "Loja Y", "agrupamento": "Fantasma"}],
    )
    stats = importar_arquivo(arq, empresa_id=e)
    assert stats["importados"] == 1  # local conhecido → entra (não pula por agrupamento)
    assert stats["linhas_agrupamento_desconhecido"] == 1
    assert db_session.query(Agrupamento).filter_by(empresa_id=e).count() == 0  # não criou
