"""Testes dos endpoints do Painel Executivo (Bloco 5 CP-1)."""

from __future__ import annotations

from datetime import datetime, timedelta

from src.models.verbatim import Verbatim


def _criar_verbatim(
    db_session,
    empresa_id,
    fonte_id,
    local_id=None,
    texto="t",
    subpilar="Pa1",
    tipo="promotor",
    data_dias_atras=10,
):
    v = Verbatim(
        empresa_id=empresa_id,
        fonte_id=fonte_id,
        local_id=local_id,
        texto=texto,
        data_criacao_original=datetime.utcnow() - timedelta(days=data_dias_atras),
        hash_dedup=f"hash-{texto}-{datetime.utcnow().timestamp()}",
        subpilar=subpilar,
        tipo=tipo,
    )
    db_session.add(v)
    db_session.commit()
    return v


def _empresa_estrutura(client_loyall, suffix=None):
    import uuid

    sfx = suffix or uuid.uuid4().hex[:6]
    e = client_loyall.post("/api/empresas/", json={"nome": f"EPnl-{sfx}"}).get_json()
    a = client_loyall.post(
        f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "GPnl"}
    ).get_json()
    loc = client_loyall.post(
        f"/api/empresas/{e['id']}/locais",
        json={"nome": "LPnl", "agrupamento_id": a["id"]},
    ).get_json()
    f = client_loyall.post(
        f"/api/locais/{loc['id']}/fontes",
        json={"conector_tipo": "google", "url": f"ChIJ_p_{sfx}"},
    ).get_json()
    return {"e": e, "a": a, "loc": loc, "f": f}


# ── Nível 1: 4 pilares ─────────────────────────────────────────────────


def test_painel_nivel1_vazio(client_loyall):
    e = client_loyall.post("/api/empresas/", json={"nome": "EPnlVazio"}).get_json()
    r = client_loyall.get(f"/api/empresas/{e['id']}/painel/nivel1")
    assert r.status_code == 200
    body = r.get_json()
    assert body["total_verbatins"] == 0
    assert len(body["pilares"]) == 4
    assert [p["pilar"] for p in body["pilares"]] == ["P", "D", "Pa", "A"]
    for p in body["pilares"]:
        assert p["total"] == 0


def test_painel_nivel1_agrega_por_pilar(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    # 2 Pa1 promotor, 1 D2 detrator, 1 P3 conversivel, 1 sem_lastro inativo
    for _ in range(2):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"t-{_}-a",
            subpilar="Pa1",
            tipo="promotor",
        )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t-d2",
        subpilar="D2",
        tipo="detrator",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t-p3",
        subpilar="P3",
        tipo="conversivel",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t-sl",
        subpilar="sem_lastro",
        tipo="inativo",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1")
    body = r.get_json()
    assert body["total_verbatins"] == 5

    pilar = {p["pilar"]: p for p in body["pilares"]}
    assert pilar["Pa"]["total"] == 2
    assert pilar["Pa"]["promotor"] == 2
    assert pilar["D"]["total"] == 1
    assert pilar["D"]["detrator"] == 1
    assert pilar["P"]["total"] == 1
    assert pilar["P"]["conversivel"] == 1
    assert pilar["A"]["total"] == 0
    assert body["outros"]["sem_lastro"] == 1


def test_painel_nivel1_filtro_periodo(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="recente",
        data_dias_atras=3,
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="velho",
        data_dias_atras=400,
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1?periodo=7d")
    assert r.get_json()["total_verbatins"] == 1
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1?periodo=12m")
    assert r.get_json()["total_verbatins"] == 1
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1")
    assert r.get_json()["total_verbatins"] == 2


def test_painel_nivel1_filtro_agrupamento(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    a2 = client_loyall.post(
        f"/api/empresas/{ctx['e']['id']}/agrupamentos", json={"nome": "G2"}
    ).get_json()
    loc2 = client_loyall.post(
        f"/api/empresas/{ctx['e']['id']}/locais",
        json={"nome": "L2", "agrupamento_id": a2["id"]},
    ).get_json()
    f2 = client_loyall.post(
        f"/api/locais/{loc2['id']}/fontes",
        json={"conector_tipo": "google", "url": "ChIJ_g2"},
    ).get_json()
    _criar_verbatim(db_session, ctx["e"]["id"], ctx["f"]["id"], ctx["loc"]["id"], texto="g1")
    _criar_verbatim(db_session, ctx["e"]["id"], f2["id"], loc2["id"], texto="g2")
    r = client_loyall.get(
        f"/api/empresas/{ctx['e']['id']}/painel/nivel1?agrupamento_id={ctx['a']['id']}"
    )
    assert r.get_json()["total_verbatins"] == 1


def test_painel_nivel1_periodo_invalido_400(client_loyall):
    e = client_loyall.post("/api/empresas/", json={"nome": "EPnlInv"}).get_json()
    r = client_loyall.get(f"/api/empresas/{e['id']}/painel/nivel1?periodo=2y")
    assert r.status_code == 400


def test_painel_nivel1_periodos_novos(client_loyall, db_session):
    """Manual Cap. 4 + pedido user: 6m e 15m disponíveis."""
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t1",
        data_dias_atras=100,
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t2",
        data_dias_atras=300,
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t3",
        data_dias_atras=500,
    )
    r6m = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1?periodo=6m")
    assert r6m.get_json()["total_verbatins"] == 1  # só os 100 dias
    r15m = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1?periodo=15m")
    assert r15m.get_json()["total_verbatins"] == 2  # 100 e 300 dias


# ── Nível 2: matriz subpilar × tipo ─────────────────────────────────────


def test_painel_nivel2_matriz_completa(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t1",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t2",
        subpilar="Pa1",
        tipo="detrator",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t3",
        subpilar="D1",
        tipo="conversivel",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel2")
    body = r.get_json()
    assert body["total_verbatins"] == 3
    assert len(body["matriz"]) == 12  # P1-3, D1-3, Pa1-3, A1-3
    pa1 = next(c for c in body["matriz"] if c["subpilar"] == "Pa1")
    assert pa1["promotor"] == 1
    assert pa1["detrator"] == 1
    assert pa1["total"] == 2
    # Manual Cap. 2: Pa1 = Empatia Comercial
    assert pa1["nome"] == "Empatia Comercial"
    # ratio = 1 promotor / 1 detrator = 1.0 → faixa "atencao"
    assert pa1["ratio"] == 1.0
    assert pa1["faixa"] == "atencao"
    d1 = next(c for c in body["matriz"] if c["subpilar"] == "D1")
    assert d1["conversivel"] == 1
    assert d1["total"] == 1
    assert d1["nome"] == "Acessibilidade"
    a1 = next(c for c in body["matriz"] if c["subpilar"] == "A1")
    assert a1["total"] == 0  # zero é coberto
    assert a1["nome"] == "Exemplo"


def test_painel_nivel2_inclui_sem_lastro_separado(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="sl",
        subpilar="sem_lastro",
        tipo="inativo",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel2")
    body = r.get_json()
    assert body["sem_lastro"]["total"] == 1
    assert body["sem_lastro"]["inativo"] == 1
    # Não polui a matriz principal
    assert all(c["total"] == 0 for c in body["matriz"])


def test_painel_nivel2_filtro_periodo_7d(client_loyall, db_session):
    """Manual Cap. 4 + hotfix: data_de/data_ate removidos, só periodo."""
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t",
        subpilar="Pa1",
        tipo="promotor",
        data_dias_atras=5,
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t2",
        subpilar="Pa1",
        tipo="promotor",
        data_dias_atras=30,
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel2?periodo=7d")
    assert r.get_json()["total_verbatins"] == 1


# ── Ratio P/D (Manual Cap. 4) ─────────────────────────────────────────


def test_calcular_ratio_zero_e_zero():
    from src.api.painel import calcular_ratio

    assert calcular_ratio(0, 0) == 0.0


def test_calcular_ratio_zero_detratores_cap_999():
    from src.api.painel import calcular_ratio

    # Saturação positiva — Manual Cap. 4
    assert calcular_ratio(10, 0) == 9.99
    assert calcular_ratio(1, 0) == 9.99


def test_calcular_ratio_zero_promotores_zero():
    from src.api.painel import calcular_ratio

    # Saturação negativa — Manual Cap. 4
    assert calcular_ratio(0, 10) == 0.0


def test_calcular_ratio_normal():
    from src.api.painel import calcular_ratio

    assert calcular_ratio(2, 1) == 2.0
    assert calcular_ratio(1, 2) == 0.5
    assert calcular_ratio(3, 1) == 3.0


def test_calcular_ratio_cap_em_caso_muito_alto():
    from src.api.painel import calcular_ratio

    # 1000 promotores, 1 detrator → 1000.0 mas capeado em 9.99
    assert calcular_ratio(1000, 1) == 9.99


def test_ratio_em_palavras_casos_do_spec():
    from src.api.painel import ratio_em_palavras

    # cap superior → sem detratores; cap inferior/zero → nenhum promotor
    assert ratio_em_palavras(9.99) == "sem detratores"
    assert ratio_em_palavras(0.0) == "nenhum promotor"
    # ratio ≥ 1: "X promotores para cada detrator" (singular quando X=1)
    assert ratio_em_palavras(6.0) == "6 promotores para cada detrator"
    assert ratio_em_palavras(2.0) == "2 promotores para cada detrator"
    assert ratio_em_palavras(1.0) == "1 promotor para cada detrator"
    # ratio < 1: "1 promotor para cada X detratores"
    assert ratio_em_palavras(0.5) == "1 promotor para cada 2 detratores"
    assert ratio_em_palavras(0.125) == "1 promotor para cada 8 detratores"


def test_ratio_em_palavras_decimal_virgula_ptbr():
    from src.api.painel import ratio_em_palavras

    assert ratio_em_palavras(1.5) == "1,5 promotores para cada detrator"
    assert ratio_em_palavras(3.33) == "3,3 promotores para cada detrator"  # 1 casa
    assert ratio_em_palavras(0.9) == "1 promotor para cada 1,1 detratores"


# ── Histórico de ratio por quarter (CP-ratio-historico) ──────────────────


def _empresa_loja(client_loyall, sfx):
    e = client_loyall.post("/api/empresas/", json={"nome": f"EHist-{sfx}"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "G"}).get_json()
    loc = client_loyall.post(
        f"/api/empresas/{e['id']}/locais", json={"nome": "L", "agrupamento_id": a["id"]}
    ).get_json()
    return e["id"], loc["id"], a["id"]


def _rm(db_session, empresa_id, local_id, ag_id, subpilar, periodo, prom, det):
    from src.api.painel import calcular_ratio
    from src.models.anomalia import RatioMensal

    db_session.add(
        RatioMensal(
            empresa_id=empresa_id,
            local_id=local_id,
            agrupamento_id=ag_id,
            subpilar=subpilar,
            periodo=periodo,
            promotor=prom,
            conversivel=0,
            detrator=det,
            total=prom + det,
            ratio=calcular_ratio(prom, det),
        )
    )


def test_historico_quarters_pondera_por_volume_e_agrega_pilar(client_loyall, db_session):
    from src.api.painel import historico_quarters_pilares

    eid, lid, agid = _empresa_loja(client_loyall, "pond")
    # Pilar P (P1+P2), Q1/2026 = jan+fev+mar; Q2/2026 = abr
    _rm(db_session, eid, lid, agid, "P1", "2026-01", 2, 8)
    _rm(db_session, eid, lid, agid, "P1", "2026-02", 1, 2)
    _rm(db_session, eid, lid, agid, "P2", "2026-03", 1, 0)
    _rm(db_session, eid, lid, agid, "P1", "2026-04", 6, 2)
    db_session.commit()

    h = historico_quarters_pilares(db_session, eid)
    # Q1: Σprom=4, Σdet=10 → 0.4 (ponderado, não média dos ratios mensais); Q2: 6/2=3.0
    assert [x["q"] for x in h["P"]] == ["Q1", "Q2"]
    assert h["P"][0]["ratio"] == 0.4 and h["P"][1]["ratio"] == 3.0


def test_historico_quarters_menos_de_2_omitido(client_loyall, db_session):
    from src.api.painel import historico_quarters_pilares

    eid, lid, agid = _empresa_loja(client_loyall, "um")
    _rm(db_session, eid, lid, agid, "P1", "2026-01", 5, 5)  # só 1 quarter
    db_session.commit()
    assert "P" not in historico_quarters_pilares(db_session, eid)


def test_historico_quarters_respeita_escopo_loja(client_loyall, db_session):
    from src.api.painel import historico_quarters_pilares

    eid, lid1, agid = _empresa_loja(client_loyall, "esc")
    loc2 = client_loyall.post(
        f"/api/empresas/{eid}/locais", json={"nome": "L2", "agrupamento_id": agid}
    ).get_json()
    lid2 = loc2["id"]
    # loja1 só pilar P; loja2 só pilar D — ambos com 2 quarters
    _rm(db_session, eid, lid1, agid, "P1", "2026-01", 3, 1)
    _rm(db_session, eid, lid1, agid, "P1", "2026-04", 4, 1)
    _rm(db_session, eid, lid2, agid, "D1", "2026-01", 2, 2)
    _rm(db_session, eid, lid2, agid, "D1", "2026-04", 1, 3)
    db_session.commit()

    h_emp = historico_quarters_pilares(db_session, eid)  # empresa-wide: agrega as 2 lojas
    assert "P" in h_emp and "D" in h_emp
    h_l1 = historico_quarters_pilares(db_session, eid, local_id=lid1)
    assert "P" in h_l1 and "D" not in h_l1  # escopo loja1 não vê D da loja2


def test_historico_quarters_ultimos_4_e_ordem(client_loyall, db_session):
    from src.api.painel import historico_quarters_pilares

    eid, lid, agid = _empresa_loja(client_loyall, "u4")
    # 5 quarters consecutivos de P1 → mantém só os 4 mais recentes, do mais antigo
    _rm(db_session, eid, lid, agid, "P1", "2025-01", 1, 1)  # Q1/25 (cai fora)
    _rm(db_session, eid, lid, agid, "P1", "2025-04", 2, 1)  # Q2/25
    _rm(db_session, eid, lid, agid, "P1", "2025-07", 3, 1)  # Q3/25
    _rm(db_session, eid, lid, agid, "P1", "2025-10", 4, 1)  # Q4/25
    _rm(db_session, eid, lid, agid, "P1", "2026-01", 5, 1)  # Q1/26
    db_session.commit()
    qs = [x["q"] for x in historico_quarters_pilares(db_session, eid)["P"]]
    assert qs == ["Q2", "Q3", "Q4", "Q1"]  # 4 últimos, antigo→recente


# ── Drawer de detalhe de quarter (CP-quarter-detalhe) ────────────────────

_QD_SEQ = [0]


def _qd_seq():
    _QD_SEQ[0] += 1
    return str(_QD_SEQ[0])


def _emp_2lojas(client_loyall):
    e = client_loyall.post("/api/empresas/", json={"nome": f"EQD-{_qd_seq()}"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "G"}).get_json()
    locs = []
    for nm in ("L1", "L2"):
        loc = client_loyall.post(
            f"/api/empresas/{e['id']}/locais", json={"nome": nm, "agrupamento_id": a["id"]}
        ).get_json()
        f = client_loyall.post(
            f"/api/locais/{loc['id']}/fontes",
            json={"conector_tipo": "google", "url": f"ChIJ_{nm}_{_qd_seq()}"},
        ).get_json()
        locs.append((loc, f))
    return e["id"], a["id"], locs


def _anom(db_session, eid, lid, agid, subpilar, periodo, severidade):
    from src.models.anomalia import AnomaliaDetectada

    db_session.add(
        AnomaliaDetectada(
            empresa_id=eid,
            local_id=lid,
            agrupamento_id=agid,
            subpilar=subpilar,
            periodo=periodo,
            severidade=severidade,
            tipo="indicador",
        )
    )


def _verb_tema(db_session, eid, lid, fid, subpilar, periodo_ym, tema_nome, n):
    from datetime import datetime

    from src.models.temas import Tema, VerbatimTema
    from src.models.verbatim import Verbatim

    slug = tema_nome.lower().replace(" ", "-")
    tema = db_session.query(Tema).filter_by(empresa_id=eid, slug=slug).first()
    if tema is None:
        tema = Tema(empresa_id=eid, nome=tema_nome, slug=slug)
        db_session.add(tema)
        db_session.flush()
    y, m = periodo_ym.split("-")
    for _ in range(n):
        v = Verbatim(
            empresa_id=eid,
            fonte_id=fid,
            local_id=lid,
            texto="x",
            subpilar=subpilar,
            tipo="detrator",
            tem_texto=True,
            data_criacao_original=datetime(int(y), int(m), 15),
            hash_dedup=f"qd{_qd_seq()}",
        )
        db_session.add(v)
        db_session.flush()
        db_session.add(VerbatimTema(verbatim_id=v.id, tema_id=tema.id, confianca=0.9, origem="llm"))


def test_quarter_detalhe_variacao_e_loja_mais_impactou(client_loyall, db_session):
    from src.api.painel import quarter_detalhe_pilar

    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, _f1), (l2, _f2) = locs
    # prev Q4/2025: ambas 1.0 (5/5)
    _rm(db_session, eid, l1["id"], agid, "P1", "2025-11", 5, 5)
    _rm(db_session, eid, l2["id"], agid, "P1", "2025-11", 5, 5)
    # sel Q1/2026: L1 despenca (0.25, vol 10 → contrib -7.5); L2 sobe (1.5, contrib +5.0)
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-02", 2, 8)
    _rm(db_session, eid, l2["id"], agid, "P1", "2026-02", 6, 4)
    db_session.commit()

    d = quarter_detalhe_pilar(db_session, eid, "P", 2026, 1)
    # pilar P: sel 8/12=0.67, prev 10/10=1.0 → -0.33 vs Q4
    assert d["variacao"]["quarter_anterior"] == "Q4"
    assert d["variacao"]["delta"] == -0.33
    # L1 impacta mais (|−7.5| > |+5.0|): ratio 0.25, variação −0.75
    assert d["loja"]["nome"] == "L1"
    assert d["loja"]["ratio"] == 0.25 and d["loja"]["variacao"] == -0.75


def test_quarter_detalhe_primeiro_quarter_sem_variacao(client_loyall, db_session):
    from src.api.painel import quarter_detalhe_pilar

    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, _f1), _ = locs
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-02", 3, 1)  # único quarter
    db_session.commit()
    d = quarter_detalhe_pilar(db_session, eid, "P", 2026, 1)
    assert d["variacao"] is None and d["loja"] is None  # 1º quarter: omite ambos


def test_quarter_detalhe_anomalia_presente_e_ausente(client_loyall, db_session):
    from src.api.painel import quarter_detalhe_pilar

    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, _f1), _ = locs
    _rm(db_session, eid, l1["id"], agid, "P1", "2025-11", 5, 5)
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-02", 2, 8)
    db_session.commit()

    assert quarter_detalhe_pilar(db_session, eid, "P", 2026, 1)["anomalia"] is None
    _anom(db_session, eid, l1["id"], agid, "P1", "2026-02", "critico")
    db_session.commit()
    a = quarter_detalhe_pilar(db_session, eid, "P", 2026, 1)["anomalia"]
    assert a["severidade"] == "critico" and "Calibração da Promessa" in a["titulo"]


def test_quarter_detalhe_tema_dominante(client_loyall, db_session):
    from src.api.painel import quarter_detalhe_pilar

    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, f1), _ = locs
    _rm(db_session, eid, l1["id"], agid, "P1", "2025-11", 5, 5)
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-02", 2, 8)
    _verb_tema(db_session, eid, l1["id"], f1["id"], "P1", "2026-02", "Atendimento lento", 4)
    _verb_tema(db_session, eid, l1["id"], f1["id"], "P1", "2026-02", "Fila grande", 2)
    db_session.commit()
    d = quarter_detalhe_pilar(db_session, eid, "P", 2026, 1)
    assert d["tema"]["nome"] == "Atendimento lento" and d["tema"]["volume"] == 4


def test_quarter_detalhe_rota_200_e_404(client_loyall, db_session):
    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, _f1), _ = locs
    _rm(db_session, eid, l1["id"], agid, "P1", "2025-11", 5, 5)
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-02", 2, 8)
    db_session.commit()
    base = f"/empresas/{eid}/explorar/painel/quarter-detalhe"
    rv = client_loyall.get(f"{base}?pilar=P&quarter=2026Q1")
    assert rv.status_code == 200
    html = rv.get_data(as_text=True)
    assert "quarter-drawer" in html and "vs Q4" in html
    # quarter malformado e pilar inválido → 404
    assert client_loyall.get(f"{base}?pilar=P&quarter=xx").status_code == 404
    assert client_loyall.get(f"{base}?pilar=Z&quarter=2026Q1").status_code == 404


def test_painel_aviso_historico_ignora_periodo_e_fonte(client_loyall, db_session):
    eid, agid, locs = _emp_2lojas(client_loyall)
    (l1, f1), _ = locs
    # histórico (2 quarters) + verbatins recentes p/ o card P existir mesmo sob período
    _rm(db_session, eid, l1["id"], agid, "P1", "2025-11", 5, 5)
    _rm(db_session, eid, l1["id"], agid, "P1", "2026-06", 2, 8)
    _verb_tema(db_session, eid, l1["id"], f1["id"], "P1", "2026-06", "T", 3)
    db_session.commit()

    aviso = "ignora filtro de período/fonte"
    base = f"/empresas/{eid}/explorar?tab=painel"
    # sem filtros → sem aviso
    assert aviso not in client_loyall.get(base).get_data(as_text=True)
    # período ativo → aviso
    assert aviso in client_loyall.get(f"{base}&periodo=30d").get_data(as_text=True)
    # fonte ativa → aviso
    assert aviso in client_loyall.get(f"{base}&fonte_id={f1['id']}").get_data(as_text=True)


def test_faixa_ratio_5_niveis():
    from src.api.painel import faixa_ratio

    assert faixa_ratio(0.0) == "critico"
    assert faixa_ratio(0.49) == "critico"
    assert faixa_ratio(0.5) == "fraco"
    assert faixa_ratio(0.99) == "fraco"
    assert faixa_ratio(1.0) == "atencao"
    assert faixa_ratio(1.99) == "atencao"
    assert faixa_ratio(2.0) == "bom"
    assert faixa_ratio(4.99) == "bom"
    assert faixa_ratio(5.0) == "excelente"
    assert faixa_ratio(9.99) == "excelente"


def test_painel_nivel1_inclui_ratio_e_faixa(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    # 4 promotor Pa1 + 1 detrator Pa1 → ratio Pa = 4.0 = "bom"
    for i in range(4):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"prom-{i}",
            subpilar="Pa1",
            tipo="promotor",
        )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="det",
        subpilar="Pa1",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1")
    body = r.get_json()
    pa = next(p for p in body["pilares"] if p["pilar"] == "Pa")
    assert pa["ratio"] == 4.0
    assert pa["faixa"] == "bom"
    # Pilar sem dados → ratio 0.0, faixa "critico"
    p = next(p for p in body["pilares"] if p["pilar"] == "P")
    assert p["ratio"] == 0.0
    assert p["faixa"] == "critico"


def test_ui_painel_link_sem_classificacao_clicavel(client_loyall, db_session):
    """B5 ext. CP-2: ao ter verbatins sem classificação, o painel
    renderiza link clicável para /verbatins?sem_classificacao=1.
    """
    ctx = _empresa_estrutura(client_loyall)
    # 1 verbatim sem_lastro e 1 sem classificação (subpilar NULL)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="sl",
        subpilar="sem_lastro",
        tipo="inativo",
    )
    v_nc = Verbatim(
        empresa_id=ctx["e"]["id"],
        fonte_id=ctx["f"]["id"],
        local_id=ctx["loc"]["id"],
        texto="nc",
        data_criacao_original=datetime.utcnow(),
        hash_dedup="hash-nc-painel",
        subpilar=None,
        tipo=None,
    )
    db_session.add(v_nc)
    db_session.commit()

    r = client_loyall.get(f"/empresas/{ctx['e']['id']}/painel")
    html = r.data.decode()
    # Link no resumo "Fora dos 4 pilares"
    assert "sem_classificacao=1" in html
    # Linha da matriz e resumo ambas devem ter o link
    occurrences = html.count("sem_classificacao=1")
    assert occurrences >= 2, f"esperava >=2 links, obtido {occurrences}"
    # Link sem_lastro continua funcionando
    assert "subpilar=sem_lastro" in html


# ── B5 ext. CP-3: Índice Geral / Previsibilidade / Concentração ──────


def test_calcular_indice_geral_sem_volume():
    from src.api.painel import calcular_indice_geral

    assert calcular_indice_geral([]) == 0.0
    assert calcular_indice_geral([{"total": 0, "ratio": 5.0, "subpilar": "P1"}]) == 0.0


def test_calcular_indice_geral_todos_saturados_da_10():
    """Hotfix opção B: todos pilares saturados (>= 5) → Índice 10."""
    from src.api.painel import calcular_indice_geral

    matriz = [
        {"total": 100, "ratio": 9.99, "subpilar": "P1", "promotor": 100, "detrator": 0},
        {"total": 100, "ratio": 9.99, "subpilar": "D1", "promotor": 100, "detrator": 0},
        {"total": 100, "ratio": 9.99, "subpilar": "Pa1", "promotor": 100, "detrator": 0},
        {"total": 100, "ratio": 9.99, "subpilar": "A1", "promotor": 100, "detrator": 0},
    ]
    pilares = [
        {"pilar": p, "ratio": 9.99, "total": 100, "promotor": 100, "detrator": 0}
        for p in ["P", "D", "Pa", "A"]
    ]
    assert calcular_indice_geral(matriz, pilares=pilares) == 10.0


def test_calcular_indice_geral_pilar_critico_puxa_pra_baixo():
    """Hotfix opção B: pilar crítico domina, mesmo com média ponderada saturada."""
    from src.api.painel import calcular_indice_geral

    # Pa saturado em 9.99 com 2000 verbatins + P crítico em 0.4 com 100 verbatins.
    # Antes (CP-3 ingênuo): ratio_medio_ponderado ≈ 9.54 × 2 = 19 → cap 10.
    # Agora (opção B): min(0.4, 9.54) × 2 = 0.80.
    matriz = [
        {"total": 100, "ratio": 0.4, "subpilar": "P1", "promotor": 2, "detrator": 5},
        {"total": 2000, "ratio": 9.99, "subpilar": "Pa1", "promotor": 2000, "detrator": 0},
    ]
    pilares = [
        {"pilar": "P", "ratio": 0.4, "total": 100, "promotor": 2, "detrator": 5},
        {"pilar": "D", "ratio": 0.0, "total": 0, "promotor": 0, "detrator": 0},
        {"pilar": "Pa", "ratio": 9.99, "total": 2000, "promotor": 2000, "detrator": 0},
        {"pilar": "A", "ratio": 0.0, "total": 0, "promotor": 0, "detrator": 0},
    ]
    indice = calcular_indice_geral(matriz, pilares=pilares)
    # min(0.4, ratio_medio_ponderado) × 2 = 0.4 × 2 = 0.8
    assert indice == 0.8


def test_calcular_indice_geral_todos_criticos_resulta_baixo():
    """Edge: todos pilares com ratio 0.3 → Índice ~0.6."""
    from src.api.painel import calcular_indice_geral

    matriz = [
        {"total": 10, "ratio": 0.3, "subpilar": "P1", "promotor": 1, "detrator": 3},
        {"total": 10, "ratio": 0.3, "subpilar": "D1", "promotor": 1, "detrator": 3},
        {"total": 10, "ratio": 0.3, "subpilar": "Pa1", "promotor": 1, "detrator": 3},
        {"total": 10, "ratio": 0.3, "subpilar": "A1", "promotor": 1, "detrator": 3},
    ]
    pilares = [
        {"pilar": p, "ratio": 0.3, "total": 10, "promotor": 1, "detrator": 3}
        for p in ["P", "D", "Pa", "A"]
    ]
    assert calcular_indice_geral(matriz, pilares=pilares) == 0.6


def test_calcular_indice_geral_sem_pilares_arg_deriva_da_matriz():
    """Fallback: quando pilares não é passado, agrega da matriz."""
    from src.api.painel import calcular_indice_geral

    matriz = [
        {"subpilar": "P1", "total": 10, "ratio": 0.5, "promotor": 1, "detrator": 2},
        {"subpilar": "Pa1", "total": 10, "ratio": 9.99, "promotor": 10, "detrator": 0},
    ]
    # min(min(P=0.5, Pa=9.99), ratio_medio) × 2 = 0.5 × 2 = 1.0
    indice = calcular_indice_geral(matriz)
    assert indice == 1.0


def test_faixa_indice_geral():
    from src.api.painel import faixa_indice_geral

    assert faixa_indice_geral(7.5) == "saudavel"
    assert faixa_indice_geral(7.0) == "saudavel"
    assert faixa_indice_geral(6.0) == "atencao"
    assert faixa_indice_geral(5.0) == "atencao"
    assert faixa_indice_geral(4.99) == "critico"
    assert faixa_indice_geral(0.0) == "critico"


def test_calcular_previsibilidade_empresa_vazia(client_loyall, db_session):
    """Edge: empresa sem verbatins → 0 (sem locais/meses pra calcular)."""
    from src.api.painel import calcular_previsibilidade
    from src.utils.db import db_session as get_session

    e = client_loyall.post("/api/empresas/", json={"nome": "EPrev0"}).get_json()
    with get_session() as s:
        prev = calcular_previsibilidade(e["id"], s, {}, pct_conversiveis=0.0)
    # Sem locais (var=0) + sem meses (vol_temporal=0) + 0% conv =
    # (1*0.4 + 1*0.3 + 0*0.3) * 100 = 70
    assert prev == 70.0


def test_calcular_previsibilidade_1_local_so(client_loyall, db_session):
    """Edge: 1 local, vários meses → var_locais=0 (precisa >= 2 locais)."""
    from src.api.painel import calcular_previsibilidade
    from src.utils.db import db_session as get_session

    ctx = _empresa_estrutura(client_loyall)
    # 5 verbatins do mesmo local, mesmo mês
    for i in range(5):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"v{i}",
            subpilar="Pa1",
            tipo="promotor",
            data_dias_atras=10,
        )
    with get_session() as s:
        prev = calcular_previsibilidade(ctx["e"]["id"], s, {}, pct_conversiveis=0.0)
    # 1 local (< 2) → var_locais=0; 1 mês (< 3) → vol_temporal=0; 0 conv
    # → (0.4 + 0.3 + 0) * 100 = 70
    assert prev == 70.0


def test_calcular_previsibilidade_lojas_uniformes_alta(client_loyall, db_session):
    """5 lojas com mesmo ratio → var_locais=0 → previsibilidade alta."""
    from src.api.painel import calcular_previsibilidade
    from src.utils.db import db_session as get_session

    e = client_loyall.post("/api/empresas/", json={"nome": "EPrevU"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "G"}).get_json()
    for i in range(5):
        loc = client_loyall.post(
            f"/api/empresas/{e['id']}/locais",
            json={"nome": f"L{i}", "agrupamento_id": a["id"]},
        ).get_json()
        f_ = client_loyall.post(
            f"/api/locais/{loc['id']}/fontes",
            json={"conector_tipo": "google", "url": f"ChIJ_pv_{i}"},
        ).get_json()
        # 5 verbatins por loja, todos Pa1/promotor (ratio=9.99 em cada)
        for j in range(5):
            _criar_verbatim(
                db_session,
                e["id"],
                f_["id"],
                loc["id"],
                texto=f"l{i}-v{j}",
                subpilar="Pa1",
                tipo="promotor",
            )
    with get_session() as s:
        prev = calcular_previsibilidade(e["id"], s, {}, pct_conversiveis=0.0)
    # 5 lojas com ratio idêntico → var_locais=0 → contribui 100% no fator 0.4
    # 0 conv → 0 no fator 0.3
    # Score >= 70 (idealmente 100 mas meses pode ser 0 se tudo no mesmo mês)
    assert prev >= 70.0


def test_calcular_previsibilidade_lojas_dispersas_reduz_score(client_loyall, db_session):
    """Lojas com ratios muito diferentes → var_locais alto → score reduz."""
    from src.api.painel import calcular_previsibilidade
    from src.utils.db import db_session as get_session

    e = client_loyall.post("/api/empresas/", json={"nome": "EPrevD"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "G"}).get_json()
    # 5 lojas: ratios variando muito (1 promotor vs 5 detratores em alguns)
    for i in range(5):
        loc = client_loyall.post(
            f"/api/empresas/{e['id']}/locais",
            json={"nome": f"L{i}", "agrupamento_id": a["id"]},
        ).get_json()
        f_ = client_loyall.post(
            f"/api/locais/{loc['id']}/fontes",
            json={"conector_tipo": "google", "url": f"ChIJ_pvd_{i}"},
        ).get_json()
        # locais pares: 5 promotor; ímpares: 5 detrator → ratios 9.99 vs 0
        tipo = "promotor" if i % 2 == 0 else "detrator"
        for j in range(5):
            _criar_verbatim(
                db_session,
                e["id"],
                f_["id"],
                loc["id"],
                texto=f"l{i}-v{j}",
                subpilar="Pa1",
                tipo=tipo,
            )
    with get_session() as s:
        prev_dispersa = calcular_previsibilidade(e["id"], s, {}, pct_conversiveis=0.0)
    # CV alto → var_locais ~1.0 → fator (1-1)*0.4 = 0
    # Score deve cair em relação ao teste uniforme
    assert prev_dispersa < 70.0


def test_painel_nivel1_inclui_indice_previsibilidade_concentracao(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    # 10 verbatins Pa1 promotor + 5 D2 detrator
    for i in range(10):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"p{i}",
            subpilar="Pa1",
            tipo="promotor",
        )
    for i in range(5):
        _criar_verbatim(
            db_session,
            ctx["e"]["id"],
            ctx["f"]["id"],
            ctx["loc"]["id"],
            texto=f"d{i}",
            subpilar="D2",
            tipo="detrator",
        )
    body = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1").get_json()
    assert "indice_geral" in body
    assert "previsibilidade" in body
    assert "concentracao_detratores" in body
    assert "indice_geral_faixa" in body
    # Concentração precisa ≥5 locais; só 1 local aqui → None
    assert body["concentracao_detratores"] is None
    assert body["concentracao_faixa"] == "indisponivel"


def test_painel_nivel1_concentracao_calcula_com_5_locais_volume_min(client_loyall, db_session):
    """Concentração: ≥5 locais com volume MÍNIMO de 5 verbatins cada (hotfix)."""
    e = client_loyall.post("/api/empresas/", json={"nome": "EConc"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "AC"}).get_json()
    # 5 locais, cada um com 5 verbatins (4 promotor + 1 detrator) → ratio 4.0
    # Necessário >= 5 verbatins/local após hotfix.
    for i in range(5):
        loc = client_loyall.post(
            f"/api/empresas/{e['id']}/locais",
            json={"nome": f"L{i}", "agrupamento_id": a["id"]},
        ).get_json()
        f_ = client_loyall.post(
            f"/api/locais/{loc['id']}/fontes",
            json={"conector_tipo": "google", "url": f"ChIJ_c_{i}"},
        ).get_json()
        for j in range(4):
            _criar_verbatim(
                db_session,
                e["id"],
                f_["id"],
                loc["id"],
                texto=f"p{i}-{j}",
                subpilar="Pa1",
                tipo="promotor",
            )
        _criar_verbatim(
            db_session,
            e["id"],
            f_["id"],
            loc["id"],
            texto=f"d{i}",
            subpilar="Pa1",
            tipo="detrator",
        )
    body = client_loyall.get(f"/api/empresas/{e['id']}/painel/nivel1").get_json()
    # 5 detratores total, 5 nas piores 5 lojas (todas iguais) → 100%
    assert body["concentracao_detratores"] == 100.0


def test_painel_nivel1_concentracao_locais_com_pouco_volume_ignorados(client_loyall, db_session):
    """Hotfix: locais com < 5 verbatins não entram no cálculo de concentração."""
    e = client_loyall.post("/api/empresas/", json={"nome": "EConcLow"}).get_json()
    a = client_loyall.post(f"/api/empresas/{e['id']}/agrupamentos", json={"nome": "AC"}).get_json()
    # 4 locais com 1 verbatim cada (volume insuficiente) + 1 local com 5 verbatins
    for i in range(5):
        loc = client_loyall.post(
            f"/api/empresas/{e['id']}/locais",
            json={"nome": f"L{i}", "agrupamento_id": a["id"]},
        ).get_json()
        f_ = client_loyall.post(
            f"/api/locais/{loc['id']}/fontes",
            json={"conector_tipo": "google", "url": f"ChIJ_cl_{i}"},
        ).get_json()
        if i == 0:
            # 5 verbatins neste local
            for j in range(5):
                _criar_verbatim(
                    db_session,
                    e["id"],
                    f_["id"],
                    loc["id"],
                    texto=f"l0-v{j}",
                    subpilar="Pa1",
                    tipo="detrator",
                )
        else:
            # apenas 1 — abaixo do mínimo
            _criar_verbatim(
                db_session,
                e["id"],
                f_["id"],
                loc["id"],
                texto=f"l{i}-v0",
                subpilar="Pa1",
                tipo="promotor",
            )
    body = client_loyall.get(f"/api/empresas/{e['id']}/painel/nivel1").get_json()
    # Apenas 1 local com volume suficiente → < 5 com volume → None.
    assert body["concentracao_detratores"] is None


def test_ui_painel_3_cards_metricas_consolidadas(client_loyall, db_session):
    """B5 ext. CP-3: 3 cards no topo do painel (Índice/Previsibilidade/Concentração)."""
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="p",
        subpilar="Pa1",
        tipo="promotor",
    )
    r = client_loyall.get(f"/empresas/{ctx['e']['id']}/painel")
    html = r.data.decode()
    assert "Índice Geral" in html
    assert "Previsibilidade" in html
    assert "Concentração (top-5)" in html  # relabel CP-LG-3 (opção i)
    assert "Desigualdade (Gini)" in html  # card Gini adicionado no CP-LG-3


def test_painel_cliente_de_outra_empresa_403(client_loyall, client_cliente_factory):
    e_a = client_loyall.post("/api/empresas/", json={"nome": "EPnlA"}).get_json()
    e_b = client_loyall.post("/api/empresas/", json={"nome": "EPnlB"}).get_json()
    cli = client_cliente_factory(e_a["id"])
    r1 = cli.get(f"/api/empresas/{e_b['id']}/painel/nivel1")
    assert r1.status_code == 403
    r2 = cli.get(f"/api/empresas/{e_b['id']}/painel/nivel2")
    assert r2.status_code == 403


def test_painel_cliente_propria_empresa_200(client_loyall, client_cliente_factory):
    e = client_loyall.post("/api/empresas/", json={"nome": "EPnlOk"}).get_json()
    cli = client_cliente_factory(e["id"])
    assert cli.get(f"/api/empresas/{e['id']}/painel/nivel1").status_code == 200
    assert cli.get(f"/api/empresas/{e['id']}/painel/nivel2").status_code == 200


# ── descrever_escopo (hotfix UI: textos contextuais) ─────────────────


def test_descrever_escopo_vazio():
    from src.api.painel import descrever_escopo

    assert descrever_escopo("BH Airport") == "geral da BH Airport"


def test_descrever_escopo_apenas_agrupamento():
    from src.api.painel import descrever_escopo

    res = descrever_escopo("BH Airport", agrupamento_nome="Aeroporto")
    assert res == "no agrupamento Aeroporto"


def test_descrever_escopo_local_supera_agrupamento():
    from src.api.painel import descrever_escopo

    # Quando ambos preenchidos, local ganha (mais específico).
    res = descrever_escopo("BH Airport", agrupamento_nome="Lojas", local_nome="Linx Confins")
    assert res == "em Linx Confins"


def test_descrever_escopo_com_fonte_amigavel():
    from src.api.painel import descrever_escopo

    res = descrever_escopo("BH Airport", local_nome="Terminal Confins", fonte_conector="google")
    assert res == "em Terminal Confins via Google Reviews"


def test_descrever_escopo_com_periodo():
    from src.api.painel import descrever_escopo

    res = descrever_escopo("BH Airport", periodo="30d")
    assert res == "geral da BH Airport nos últimos 30 dias"


def test_descrever_escopo_combinacao_completa():
    from src.api.painel import descrever_escopo

    res = descrever_escopo(
        "BH Airport",
        local_nome="Terminal Confins",
        fonte_conector="google",
        periodo="7d",
    )
    assert res == "em Terminal Confins via Google Reviews nos últimos 7 dias"


def test_descrever_escopo_fonte_sem_local_nem_agrupamento():
    from src.api.painel import descrever_escopo

    res = descrever_escopo("BH Airport", fonte_conector="instagram")
    assert res == "geral da BH Airport via Instagram"


def test_descrever_escopo_periodo_invalido_e_ignorado():
    from src.api.painel import descrever_escopo

    res = descrever_escopo("BH Airport", periodo="2y")
    assert res == "geral da BH Airport"


def test_painel_nivel1_inclui_filtros_descricao(client_loyall, db_session):
    """Endpoint nivel1 deve devolver `filtros_descricao` no JSON."""
    ctx = _empresa_estrutura(client_loyall, "fd")
    body = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/nivel1").get_json()
    assert "filtros_descricao" in body
    # Sem filtros → "geral da <nome>"
    assert body["filtros_descricao"].startswith("geral da")


def test_painel_nivel1_filtros_descricao_com_agrupamento(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall, "fdag")
    body = client_loyall.get(
        f"/api/empresas/{ctx['e']['id']}/painel/nivel1?agrupamento_id={ctx['a']['id']}"
    ).get_json()
    assert body["filtros_descricao"] == "no agrupamento GPnl"


def test_painel_nivel1_filtros_descricao_com_local_e_periodo(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall, "fdlp")
    body = client_loyall.get(
        f"/api/empresas/{ctx['e']['id']}/painel/nivel1" f"?local_id={ctx['loc']['id']}&periodo=30d"
    ).get_json()
    assert body["filtros_descricao"] == "em LPnl nos últimos 30 dias"


# ── UI: página /empresas/<id>/painel ──────────────────────────────────


def test_ui_painel_renderiza_visao_geral_e_detalhamento(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="elogio",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="reclamacao",
        subpilar="D2",
        tipo="detrator",
    )
    r = client_loyall.get(f"/empresas/{ctx['e']['id']}/painel")
    assert r.status_code == 200
    html = r.data.decode()
    assert "Visão Geral" in html
    assert "Detalhamento por Subpilar" in html
    # Pilares P/D/Pa/A presentes
    for pilar in ["Precisão", "Disponibilidade", "Parceria", "Aconselhamento"]:
        assert pilar in html
    # Nomes oficiais dos 12 subpilares (Manual Cap. 2)
    assert "Empatia Comercial" in html
    assert "Calibração da Promessa" in html
    assert "Eficácia Operacional" in html
    # Ratio P/D presente no card de pilar e na matriz
    assert "Ratio P/D" in html
    # Períodos do manual + adicionais
    for p in ["7d", "30d", "90d", "6m", "12m", "15m"]:
        assert f'value="{p}"' in html
    # Botão exportar
    assert "Exportar Excel" in html


def test_ui_painel_403_cliente_de_outra_empresa(client_loyall, client_cliente_factory):
    e_a = client_loyall.post("/api/empresas/", json={"nome": "EPnlUiA"}).get_json()
    e_b = client_loyall.post("/api/empresas/", json={"nome": "EPnlUiB"}).get_json()
    cli = client_cliente_factory(e_a["id"])
    r = cli.get(f"/empresas/{e_b['id']}/painel")
    assert r.status_code == 403


# ── Exportar XLSX ─────────────────────────────────────────────────────


def test_exportar_painel_xlsx_basico(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t1",
        subpilar="Pa1",
        tipo="promotor",
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="t2",
        subpilar="D2",
        tipo="detrator",
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/exportar.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    assert r.data[:2] == b"PK"

    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    assert "Visão Geral" in wb.sheetnames
    assert "Detalhamento por Subpilar" in wb.sheetnames

    ws1 = wb["Visão Geral"]
    rows1 = list(ws1.iter_rows(values_only=True))
    # Confere headers da tabela de pilares (com ratio + faixa)
    header_row = next(r for r in rows1 if r and r[0] == "Pilar")
    assert header_row[0:3] == ("Pilar", "Nome", "Total")
    assert "Ratio P/D" in header_row
    assert "Faixa" in header_row
    # Pa deve aparecer com total=1, nome correto, ratio 9.99 (1 prom, 0 detr)
    pa_row = next(r for r in rows1 if r and r[0] == "Pa")
    assert pa_row[1] == "Parceria"
    assert pa_row[2] == 1
    assert pa_row[3] == 1  # promotor
    assert pa_row[7] == 9.99  # Ratio P/D (saturação positiva)
    assert pa_row[8] == "excelente"

    ws2 = wb["Detalhamento por Subpilar"]
    rows2 = list(ws2.iter_rows(values_only=True))
    header2 = next(r for r in rows2 if r and r[0] == "Pilar")
    assert "Nome do Subpilar" in header2
    assert "Ratio P/D" in header2
    pa1_row = next(r for r in rows2 if r and r[1] == "Pa1")
    assert pa1_row[2] == "Empatia Comercial"  # nome oficial
    assert pa1_row[3] == 1  # promotor (1ª coluna de tipo após Nome)
    assert pa1_row[7] == 1  # total
    assert pa1_row[8] == 9.99  # ratio
    d2_row = next(r for r in rows2 if r and r[1] == "D2")
    assert d2_row[2] == "Eficácia Operacional"
    assert d2_row[5] == 1  # detrator


def test_exportar_painel_xlsx_aplica_filtros(client_loyall, db_session):
    ctx = _empresa_estrutura(client_loyall)
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="recente",
        subpilar="Pa1",
        tipo="promotor",
        data_dias_atras=3,
    )
    _criar_verbatim(
        db_session,
        ctx["e"]["id"],
        ctx["f"]["id"],
        ctx["loc"]["id"],
        texto="velho",
        subpilar="D2",
        tipo="detrator",
        data_dias_atras=400,
    )
    r = client_loyall.get(f"/api/empresas/{ctx['e']['id']}/painel/exportar.xlsx?periodo=7d")
    assert r.status_code == 200
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.data))
    ws1 = wb["Visão Geral"]
    rows1 = list(ws1.iter_rows(values_only=True))
    pa_row = next(r for r in rows1 if r and r[0] == "Pa")
    assert pa_row[2] == 1
    d_row = next(r for r in rows1 if r and r[0] == "D")
    assert d_row[2] == 0  # filtrado pelo periodo


def test_exportar_painel_xlsx_cliente_outra_empresa_403(client_loyall, client_cliente_factory):
    e_a = client_loyall.post("/api/empresas/", json={"nome": "EPnlXlsxA"}).get_json()
    e_b = client_loyall.post("/api/empresas/", json={"nome": "EPnlXlsxB"}).get_json()
    cli = client_cliente_factory(e_a["id"])
    r = cli.get(f"/api/empresas/{e_b['id']}/painel/exportar.xlsx")
    assert r.status_code == 403
