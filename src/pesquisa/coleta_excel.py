"""Canal Excel de RESPOSTAS (Fase 2 · Passo 3a) — parser WIDE.

Diferente do import de verbatins (long: 1 linha = 1 verbatim): aqui é WIDE —
1 linha = 1 respondente, 1 coluna = 1 pergunta. Keyed a uma ``pesquisa_id``
escolhida na tela. Reusa o núcleo ``registrar_respostas`` (coleta.py) e os helpers
de células/identidade de ``excel.py``. NÃO toca o import de verbatins.

Mapeamento por ORDEM (W2): o modelo é gerado por nós e cada header de pergunta
carrega o prefixo ``P<ordem>`` (``P<ordem>n``/``P<ordem>t`` p/ as 2 colunas da
mista) → o parser extrai a ordem e resolve ``pergunta_id`` via ``pesquisa.perguntas``.
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd

from src.coletor.excel import (
    _ALIASES_IDENTIDADE,
    _ler_dataframe,
    _norm_email,
    _norm_nome,
    _parse_data,
    _parse_rating,
    _reconciliar_pessoa,
    _texto_celula,
)
from src.models.pesquisa import Pesquisa
from src.pesquisa.coleta import registrar_respostas
from src.pesquisa.persistencia import _opcoes_publicas
from src.utils.db import db_session

_RE_HEADER_PERGUNTA = re.compile(r"^p(\d+)([nt]?)\b")

# Nome de exibição (opcional) — só rótulo, NÃO chave de fusão (fusão é email/id_cliente).
# Detecção isolada aqui (não mexe nos aliases do detector de Verbatins).
_ALIASES_NOME = {"nome", "name", "nome_cliente", "nome cliente"}


def _perguntas_conteudo(pesquisa: Pesquisa) -> List[Any]:
    return [p for p in pesquisa.perguntas if not p.gerada_por_ancora]


def _ancora(pesquisa: Pesquisa) -> Optional[Any]:
    return next((p for p in pesquisa.perguntas if p.gerada_por_ancora), None)


def _header_pergunta(ordem: int, enunciado: str, sufixo: str = "") -> str:
    """Header do modelo: ``P<ordem>[n|t]. <enunciado>`` (sufixo só na mista)."""
    return f"P{ordem}{sufixo}. {enunciado}"


def gerar_modelo_respostas_xlsx(pesquisa: Pesquisa) -> io.BytesIO:
    """Modelo .xlsx das respostas de UMA pesquisa: colunas de identidade (se não
    anônima) + unidade (modo geral) + uma/duas colunas por pergunta (mista = 2).
    Headers com prefixo de ordem → o parser remapeia por ordem."""
    cols: List[str] = []
    exemplo: Dict[str, Any] = {}
    if not pesquisa.anonima:
        # nome = rótulo de exibição (opcional) — traz o nome de uma pessoa NOVA. NÃO é chave
        # de fusão (fusão só por email/id_cliente); primeiro nome vence no reconciliador.
        cols += ["email", "id_cliente", "nome"]
        exemplo["email"] = "maria.souza@empresa.com"
        exemplo["id_cliente"] = "CRM-1001"
        exemplo["nome"] = "Maria Souza"
    # Data da resposta: quando o respondente respondeu (histórico). SEM ela, todo o
    # import cai no mês do upload e a série temporal mente. Vazia → data do upload.
    cols.append("data_resposta (AAAA-MM-DD)")
    exemplo["data_resposta (AAAA-MM-DD)"] = "2026-05-10"
    unidade_col: Optional[str] = None
    rotulos_unidade: List[str] = []
    if pesquisa.escopo_local_modo == "geral":
        unidade_col = "Unidade"
        cols.append(unidade_col)
        pub = _opcoes_publicas((_ancora(pesquisa) or _Vazio()).opcoes_json)
        rotulos_unidade = pub["rotulos"] if pub and pub.get("rotulos") else []
        exemplo[unidade_col] = rotulos_unidade[0] if rotulos_unidade else "Nome da unidade"
    notas_cols: List[str] = []  # colunas *n. (e fechada) → trava 1-5
    for p in _perguntas_conteudo(pesquisa):
        if p.formato == "mista":
            hn = _header_pergunta(p.ordem, f"{p.enunciado} — nota", "n")
            ht = _header_pergunta(p.ordem, f"{p.enunciado} — comentário", "t")
            cols += [hn, ht]
            exemplo[hn] = 5
            exemplo[ht] = "Comentário de exemplo"
            notas_cols.append(hn)
        else:
            h = _header_pergunta(p.ordem, p.enunciado)
            cols.append(h)
            if p.formato == "fechada":
                exemplo[h] = 5
                notas_cols.append(h)
            else:
                exemplo[h] = "Resposta de exemplo"
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame([exemplo], columns=cols).to_excel(writer, index=False, sheet_name="respostas")
        _aplicar_validacoes_respostas(writer, cols, notas_cols, unidade_col, rotulos_unidade)
    bio.seek(0)
    return bio


def _aplicar_validacoes_respostas(writer, cols, notas_cols, unidade_col, rotulos_unidade):
    """Cercas de UX nativas do Excel no modelo de respostas (padrão herdado do Importar
    Verbatins): cada coluna de NOTA trava 1–5 (inteiro); a coluna Unidade vira DROPDOWN
    FECHADO dos rótulos da âncora, numa aba oculta 'listas'. Só o gerador — o parser ignora
    (a nota já é validada por _nota_valida). Sem rótulos → sem dropdown (empresa-wide)."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = writer.sheets["respostas"]
    book = writer.book
    ultima = 1000

    def _faixa(col: str) -> str:
        letra = get_column_letter(cols.index(col) + 1)
        return f"{letra}2:{letra}{ultima}"

    if notas_cols:  # dropdown 1-5 nas colunas de nota; *t. (comentário) ficam livres
        # Lista INLINE (não referência a aba) → validação PADRÃO (não x14) e o dropdown
        # renderiza inclusive no Numbers. allow_blank: linha sem essa nota é permitida.
        dv_n = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
        dv_n.error, dv_n.errorTitle = "A nota deve ser um inteiro de 1 a 5.", "Nota inválida"
        dv_n.showErrorMessage = True
        ws.add_data_validation(dv_n)
        for c in notas_cols:
            dv_n.add(_faixa(c))

    if unidade_col and rotulos_unidade:  # dropdown fechado das unidades da pesquisa
        lst = book.create_sheet("listas")
        lst.sheet_state = "hidden"
        for i, rot in enumerate(rotulos_unidade, start=1):
            lst[f"A{i}"] = rot
        dv_u = DataValidation(
            type="list", formula1=f"listas!$A$1:$A${len(rotulos_unidade)}", allow_blank=True
        )
        dv_u.error = "Escolha uma unidade da lista (as unidades desta pesquisa)."
        dv_u.errorTitle, dv_u.showErrorMessage = "Unidade não listada", True
        ws.add_data_validation(dv_u)
        dv_u.add(_faixa(unidade_col))


class _Vazio:
    opcoes_json = None


def _classificar_colunas(columns: List[str], ordens_validas: set) -> Dict[str, Dict[str, Any]]:
    """col → {tipo: 'pergunta'|'email'|'id_cliente'|'unidade'|None, ordem?, kind?}."""
    out: Dict[str, Dict[str, Any]] = {}
    for col in columns:
        norm = str(col).strip().lower()
        m = _RE_HEADER_PERGUNTA.match(norm)
        if m and int(m.group(1)) in ordens_validas:
            out[col] = {"tipo": "pergunta", "ordem": int(m.group(1)), "kind": m.group(2)}
        elif norm in _ALIASES_IDENTIDADE["email"]:
            out[col] = {"tipo": "email"}
        elif norm in _ALIASES_IDENTIDADE["id_cliente"]:
            out[col] = {"tipo": "id_cliente"}
        elif norm in _ALIASES_NOME:
            out[col] = {"tipo": "nome"}
        elif norm in ("unidade", "loja", "local"):
            out[col] = {"tipo": "unidade"}
        elif norm.startswith("data"):  # 'data', 'data_resposta', 'data_resposta (aaaa-mm-dd)'
            out[col] = {"tipo": "data"}
        else:
            out[col] = {"tipo": None}
    return out


def _data_resposta(row, col_class) -> Optional[datetime]:
    """Lê a coluna de data (quando o respondente respondeu). Ausente/inválida → None
    (o núcleo cai no fallback: data do upload)."""
    for col, c in col_class.items():
        if c["tipo"] == "data":
            return _parse_data(row[col])
    return None


def _escopo_por_rotulo(ancora: Optional[Any]) -> Dict[str, Tuple[str, Optional[int]]]:
    if ancora is None or not ancora.opcoes_json:
        return {}
    pub = _opcoes_publicas(ancora.opcoes_json)
    out: Dict[str, Tuple[str, Optional[int]]] = {}
    for o in (pub or {}).get("opcoes", []):
        out[str(o["rotulo"]).strip().lower()] = (o["entidade_tipo"], o["entidade_id"])
    return out


def _resolver_escopo(row, col_class, pesquisa, escopo_por_rotulo):
    if pesquisa.escopo_local_modo == "geral":
        for col, c in col_class.items():
            if c["tipo"] == "unidade":
                rot = (_texto_celula(row[col]) or "").strip().lower()
                if rot in escopo_por_rotulo:
                    return escopo_por_rotulo[rot]
    return (pesquisa.entidade_tipo or "empresa", pesquisa.entidade_id)


def _montar_respostas(row, col_class, ordem_map) -> List[Dict[str, Any]]:
    """Acumula as colunas de cada pergunta em uma entrada {pergunta_id, ...}."""
    por_pergunta: Dict[int, Dict[str, Any]] = {}
    for col, c in col_class.items():
        if c["tipo"] != "pergunta":
            continue
        p = ordem_map[c["ordem"]]
        acc = por_pergunta.setdefault(p.id, {"pergunta_id": p.id})
        kind = c["kind"]
        if kind == "n" or (kind == "" and p.formato == "fechada" and _eh_nota(p)):
            nota = _parse_rating(row[col])
            if nota is not None and _nota_valida(p, nota):
                acc["nota"] = nota
        elif kind == "t" or (kind == "" and p.formato in ("aberta", "mista")):
            txt = _texto_celula(row[col]) or None
            if txt:
                acc["texto"] = txt
        else:  # fechada-múltipla
            op = _texto_celula(row[col]) or None
            if op:
                acc["opcao"] = op
    # só perguntas com algum valor de fato
    return [v for v in por_pergunta.values() if len(v) > 1]


def _eh_nota(p) -> bool:
    import json

    if not p.opcoes_json:
        return False
    try:
        return json.loads(p.opcoes_json).get("tipo") == "nota"
    except (ValueError, TypeError):
        return False


def _nota_valida(p, nota: int) -> bool:
    import json

    try:
        pontos = json.loads(p.opcoes_json).get("pontos")
    except (ValueError, TypeError, AttributeError):
        pontos = None
    if not isinstance(pontos, int):
        return nota >= 1
    return 1 <= nota <= pontos


def _resolver_identidade(s, row, col_class, pesquisa, cache, stats):
    if pesquisa.anonima:
        return None
    email = idc = nome = None
    for col, c in col_class.items():
        if c["tipo"] == "email":
            email = _norm_email(row[col])
        elif c["tipo"] == "id_cliente":
            idc = _norm_nome(row[col])
        elif c["tipo"] == "nome":
            nome = _norm_nome(row[col])
    if not (email or idc):  # nome sozinho NÃO identifica nem cria (só rótulo)
        stats["sem_identidade"] += 1
        return None
    # Item A: e-mail E código do CRM viram chaves da MESMA Pessoa. Nome = rótulo de
    # exibição (primeiro nome vence — o reconciliador só preenche Pessoa sem nome).
    return _reconciliar_pessoa(
        s, email=email, id_cliente=idc, nome=nome, origem="import_excel_respostas"
    )


def importar_respostas(
    caminho: Union[str, Path],
    pesquisa_id: int,
    *,
    consentimento: bool = False,
    autor_id: Optional[int] = None,
    arquivo_nome: Optional[str] = None,
) -> Dict[str, Any]:
    """Importa um Excel WIDE de respostas para a pesquisa ``pesquisa_id``. Cada
    linha → 1 Respondente + suas respostas (destino pelo propósito da pesquisa)."""
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    df = _ler_dataframe(caminho)
    stats: Dict[str, Any] = {
        "respondentes": 0,
        "respostas": 0,
        "ignorados": 0,
        "erros": 0,
        "sem_identidade": 0,
        "total": len(df),
    }

    with db_session() as s:
        pesq = s.get(Pesquisa, pesquisa_id)
        if pesq is None or pesq.status != "pronta":
            return {**stats, "erros_validacao": ["Pesquisa não encontrada ou não publicada."]}
        if not pesq.anonima and not consentimento:
            return {
                **stats,
                "erros_validacao": ["Consentimento obrigatório para respostas identificadas."],
            }

        perguntas = _perguntas_conteudo(pesq)
        ordem_map = {p.ordem: p for p in perguntas}
        col_class = _classificar_colunas(list(df.columns), set(ordem_map))
        if not any(c["tipo"] == "pergunta" for c in col_class.values()):
            return {
                **stats,
                "erros_validacao": ["Nenhuma coluna de pergunta reconhecida — use o modelo."],
                "colunas_detectadas": col_class,
            }
        escopo_por_rotulo = _escopo_por_rotulo(_ancora(pesq))

        # Onda 2: lote desfazível — carimba cada Respondente/Verbatim deste import.
        from src.models.importacao import ImportacaoLote

        lote = ImportacaoLote(
            empresa_id=pesq.empresa_id,
            tipo="respostas",
            arquivo_nome=arquivo_nome or Path(caminho).name,
            autor_id=autor_id,
        )
        s.add(lote)
        s.flush()

        cache_pessoa: Dict[str, int] = {}
        for _, row in df.iterrows():
            try:
                respostas = _montar_respostas(row, col_class, ordem_map)
                if not respostas:
                    stats["ignorados"] += 1  # linha sem nenhuma resposta
                    continue
                pessoa_id = _resolver_identidade(s, row, col_class, pesq, cache_pessoa, stats)
                escopo = _resolver_escopo(row, col_class, pesq, escopo_por_rotulo)
                registrar_respostas(
                    s,
                    pesq,
                    escopo=escopo,
                    pessoa_id=pessoa_id,
                    respostas=respostas,
                    conector="pesquisa_excel",
                    data_resposta=_data_resposta(row, col_class),  # da planilha, não do upload
                    lote_id=lote.id,
                )
                stats["respondentes"] += 1
                stats["respostas"] += len(respostas)
            except Exception:  # noqa: BLE001 — linha problemática não derruba o lote
                stats["erros"] += 1

        lote.contadores_json = json.dumps(
            {k: stats[k] for k in ("respondentes", "respostas", "ignorados", "erros", "total")}
        )
        stats["lote_id"] = lote.id

    return stats
