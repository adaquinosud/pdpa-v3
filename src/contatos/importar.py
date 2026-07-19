"""Import da base de contatos — modelo .xlsx, preview e UPSERT.

Colunas FIXAS: email, id_cliente, nome, unidade. Todo o RESTO do arquivo é
candidato a ATRIBUTO — o operador marca no mapeamento quais viram atributo (default:
tudo desmarcado, LGPD). O import é UPSERT: cria quem falta, atualiza quem existe,
NUNCA apaga; ausentes ficam intocados (ou marcados ``inativo`` se o operador disser
que é a base completa).

A PORTA DE CRIAÇÃO do "convidado" reusa ``_reconciliar_pessoa`` (mesma chave
email→'pesquisa' / id_cliente→'crm'), sem duplicar a reconciliação — assim o contato
COLAPSA com a mesma Pessoa quando responder. ``importar_contatos`` é o CHOKE POINT
único de escrita (a porta que a Onda 2 instrumenta com ``import_id``).
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd

from src.coletor.excel import (
    FONTE_CRM,
    FONTE_EMAIL,
    _find_or_create_local,
    _ler_dataframe,
    _norm_email,
    _norm_nome,
    _reconciliar_pessoa,
)
from src.contatos.atributos import termo_sensivel, upsert_atributo
from src.models.contato import ContatoEmpresa
from src.models.pessoa import PessoaIdentificador

# Campo lógico fixo → aliases (nome inteiro normalizado OU token do cabeçalho).
_ALIASES_FIXAS: Dict[str, set[str]] = {
    "email": {"email", "e-mail", "e_mail", "mail"},
    "id_cliente": {"id_cliente", "idcliente", "codigo", "cod", "crm", "matricula"},
    "nome": {"nome", "name", "contato"},
    "unidade": {"unidade", "loja", "local", "filial"},
}
_COLS_FIXAS = list(_ALIASES_FIXAS)


def _detectar_fixas(columns: List[str]) -> tuple[Dict[str, Optional[str]], List[str]]:
    """Mapeia campo fixo → nome real da coluna (ou None) e retorna as colunas EXTRAS
    (candidatas a atributo). Cada coluna vai p/ no máx. 1 campo fixo; o resto é extra."""
    mapping: Dict[str, Optional[str]] = {k: None for k in _ALIASES_FIXAS}
    usados: set[int] = set()
    for campo, aliases in _ALIASES_FIXAS.items():
        for idx, col in enumerate(columns):
            if idx in usados:
                continue
            norm = str(col).strip().lower()
            tokens = {t for t in re.split(r"[^a-z0-9]+", norm) if t}
            if norm in aliases or (tokens & aliases):
                mapping[campo] = col
                usados.add(idx)
                break
    extras = [str(col) for idx, col in enumerate(columns) if idx not in usados]
    return mapping, extras


def gerar_modelo_contatos_xlsx(locais: Optional[List[str]] = None) -> io.BytesIO:
    """Modelo .xlsx da base de contatos: colunas fixas (email, id_cliente, nome,
    unidade) + colunas de EXEMPLO de atributo (plano, cidade) que o operador marca no
    import. ``unidade`` vira DROPDOWN FECHADO da lista de locais da empresa (padrão do
    modelo de verbatins) — não-casado nunca cria unidade nova."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    cols = _COLS_FIXAS + ["plano", "cidade"]
    rows = [
        {
            "email": "maria.souza@empresa.com",
            "id_cliente": "CRM-1001",
            "nome": "Maria Souza",
            "unidade": "Loja Centro",
            "plano": "premium",
            "cidade": "Belo Horizonte",
        },
        {
            "email": "joao.lima@empresa.com",
            "id_cliente": "CRM-1002",
            "nome": "João Lima",
            "unidade": "Loja Shopping",
            "plano": "básico",
            "cidade": "Contagem",
        },
    ]
    instr = [
        ("email", "E-mail do contato — CHAVE de identidade (colapsa com a resposta)."),
        ("id_cliente", "Código do cliente no CRM/ERP — CHAVE de identidade."),
        ("nome", "Nome do contato (rótulo — NUNCA identifica sozinho)."),
        ("unidade", "A loja/filial do contato (dropdown da lista da empresa)."),
        ("plano", "EXEMPLO de atributo — vira atributo SÓ se você marcar no import."),
        ("cidade", "EXEMPLO de atributo — idem. Toda coluna extra é candidata."),
    ]
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=cols).to_excel(writer, index=False, sheet_name="contatos")
        pd.DataFrame(instr, columns=["coluna", "o que preencher"]).to_excel(
            writer, index=False, sheet_name="instruções"
        )
        locais = locais or []
        if locais:
            ws = writer.sheets["contatos"]
            book = writer.book
            lst = book.create_sheet("listas")
            lst.sheet_state = "hidden"
            for i, nome in enumerate(locais, start=1):
                lst[f"A{i}"] = nome
            letra = get_column_letter(cols.index("unidade") + 1)
            dv = DataValidation(
                type="list", formula1=f"listas!$A$1:$A${len(locais)}", allow_blank=True
            )
            dv.error = "Escolha uma unidade da lista (cadastre-a na empresa antes)."
            dv.errorTitle, dv.showErrorMessage = "Unidade não cadastrada", True
            ws.add_data_validation(dv)
            dv.add(f"{letra}2:{letra}1000")
    bio.seek(0)
    return bio


def _chave_linha(row: pd.Series, fixas: Dict[str, Optional[str]]) -> tuple:
    """Extrai (email, id_cliente, nome) normalizados de uma linha."""
    email = _norm_email(row[fixas["email"]]) if fixas["email"] else None
    id_cliente = _norm_nome(row[fixas["id_cliente"]]) if fixas["id_cliente"] else None
    nome = _norm_nome(row[fixas["nome"]]) if fixas["nome"] else None
    return email, id_cliente, nome


def _pessoa_por_chave(session, email: Optional[str], id_cliente: Optional[str]) -> Optional[int]:
    """Lookup READ-ONLY da Pessoa por qualquer chave (sem criar). Usado no preview."""
    for fonte, ext in ((FONTE_EMAIL, email), (FONTE_CRM, id_cliente)):
        if not ext:
            continue
        ident = (
            session.query(PessoaIdentificador)
            .filter_by(tipo="interno_consentido", fonte=fonte, external_id=ext)
            .first()
        )
        if ident is not None:
            return ident.pessoa_id
    return None


def prever_contatos(session, caminho: Union[str, Path], empresa_id: int) -> Dict[str, Any]:
    """Preview (read-only, sem escrever): detecta colunas fixas + extras, valida que
    exista pelo menos uma coluna-chave, e conta criar/atualizar/ignorar-sem-chave +
    avisos de coluna sensível. Estilo ``coletor.excel.prever_arquivo``, mas com lookup
    no banco p/ separar criar (novo vínculo) de atualizar (já é contato)."""
    df = _ler_dataframe(Path(caminho))
    fixas, extras = _detectar_fixas(list(df.columns))
    erros: List[str] = []
    if fixas["email"] is None and fixas["id_cliente"] is None:
        erros.append("Nenhuma coluna de e-mail nem de id_cliente (precisa de ao menos uma chave).")

    avisos_sensiveis = [
        {"coluna": col, "categoria": cat}
        for col in extras
        if (cat := termo_sensivel(col)) is not None
    ]

    criar = atualizar = ignorar = 0
    if not erros:
        for _, row in df.iterrows():
            email, id_cliente, _nome = _chave_linha(row, fixas)
            if not email and not id_cliente:
                ignorar += 1
                continue
            pessoa_id = _pessoa_por_chave(session, email, id_cliente)
            ja_contato = pessoa_id is not None and (
                session.query(ContatoEmpresa)
                .filter_by(empresa_id=empresa_id, pessoa_id=pessoa_id)
                .first()
                is not None
            )
            atualizar += 1 if ja_contato else 0
            criar += 0 if ja_contato else 1

    return {
        "total": len(df),
        "headers": [str(c) for c in df.columns],
        "fixas": fixas,
        "extras": extras,
        "erros": erros,
        "avisos_sensiveis": avisos_sensiveis,
        "criar": criar,
        "atualizar": atualizar,
        "ignorar_sem_chave": ignorar,
    }


def _convidar(
    session,
    empresa_id: int,
    *,
    email: Optional[str],
    id_cliente: Optional[str],
    nome: Optional[str],
    local_id: Optional[int],
) -> Optional[int]:
    """Cria/reusa a Pessoa "convidado" pela chave (REUSA ``_reconciliar_pessoa``) e
    faz upsert do vínculo por-empresa. Sem chave → None (linha ignorada). Import é
    UPSERT: presente no arquivo = ``ativo`` (reativa contato antes inativado)."""
    pessoa_id = _reconciliar_pessoa(
        session, email=email, id_cliente=id_cliente, nome=nome, origem="contato"
    )
    if pessoa_id is None:
        return None
    vinculo = (
        session.query(ContatoEmpresa).filter_by(empresa_id=empresa_id, pessoa_id=pessoa_id).first()
    )
    if vinculo is None:
        session.add(
            ContatoEmpresa(
                empresa_id=empresa_id, pessoa_id=pessoa_id, local_id=local_id, status="ativo"
            )
        )
    else:
        if local_id is not None:
            vinculo.local_id = local_id
        vinculo.status = "ativo"
    return pessoa_id


def importar_contatos(
    session,
    caminho: Union[str, Path],
    empresa_id: int,
    *,
    atributos_marcados: Optional[List[str]] = None,
    marcar_ausentes_inativo: bool = False,
) -> Dict[str, Any]:
    """CHOKE POINT do import de contatos (UPSERT). Para cada linha com chave: convida
    (Pessoa + vínculo) e grava os atributos MARCADOS presentes. ``atributos_marcados``
    default vazio = nada vira atributo. ``marcar_ausentes_inativo`` (checkbox opcional
    "base completa") marca ``inativo`` os contatos ATIVOS não tocados — NUNCA apaga."""
    atributos_marcados = atributos_marcados or []
    df = _ler_dataframe(Path(caminho))
    fixas, _extras = _detectar_fixas(list(df.columns))
    marcados = [c for c in atributos_marcados if c in df.columns]

    cache_local: Dict[str, int] = {}
    tocados: set[int] = set()
    stats = {
        "criados": 0,
        "atualizados": 0,
        "ignorados_sem_chave": 0,
        "atributos_gravados": 0,
        "unidades_nao_casadas": 0,
        "inativados": 0,
    }
    for _, row in df.iterrows():
        email, id_cliente, nome = _chave_linha(row, fixas)
        if not email and not id_cliente:
            stats["ignorados_sem_chave"] += 1
            continue

        local_id = None
        if fixas["unidade"]:
            uni = _norm_nome(row[fixas["unidade"]])
            if uni:
                local_id = _find_or_create_local(
                    session, empresa_id, uni, None, cache_local, criar=False
                )
                if local_id is None:
                    stats["unidades_nao_casadas"] += 1

        pid_existente = _pessoa_por_chave(session, email, id_cliente)
        novo = pid_existente is None or (
            session.query(ContatoEmpresa)
            .filter_by(empresa_id=empresa_id, pessoa_id=pid_existente)
            .first()
            is None
        )

        pessoa_id = _convidar(
            session,
            empresa_id,
            email=email,
            id_cliente=id_cliente,
            nome=nome,
            local_id=local_id,
        )
        if pessoa_id is None:
            stats["ignorados_sem_chave"] += 1
            continue
        tocados.add(pessoa_id)
        stats["criados" if novo else "atualizados"] += 1

        for col in marcados:
            desfecho = upsert_atributo(session, empresa_id, pessoa_id, col, row[col])
            if desfecho in ("criado", "mudou"):
                stats["atributos_gravados"] += 1

    session.flush()
    if marcar_ausentes_inativo:
        ausentes = (
            session.query(ContatoEmpresa)
            .filter(
                ContatoEmpresa.empresa_id == empresa_id,
                ContatoEmpresa.status == "ativo",
                ~ContatoEmpresa.pessoa_id.in_(tocados) if tocados else True,
            )
            .all()
        )
        for v in ausentes:
            v.status = "inativo"
        stats["inativados"] = len(ausentes)
    return stats
