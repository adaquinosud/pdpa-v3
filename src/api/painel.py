"""Painel Executivo (Bloco 5) — endpoints de agregação por pilar/subpilar.

Dois endpoints expostos via blueprint, registrados sob ``/api/empresas``:

- ``GET /api/empresas/<id>/painel/nivel1`` — totais por pilar (P, D, Pa, A).
- ``GET /api/empresas/<id>/painel/nivel2`` — matriz subpilar × tipo (12×3).

Decisão arquitetural CP1: **runtime, sem materialização**. SQLite resolve
``GROUP BY subpilar, tipo`` em ms para volumes < 500k verbatins. Quando o
volume justificar, materializamos via job + tabela ``painel_snapshot``
(pendência registrada em PENDENCIAS_TECNICAS.md).

Filtros aceitos via query string:

- ``agrupamento_id`` (int) — restringe a verbatins de locais do agrupamento
- ``local_id`` (int) — restringe a um local específico
- ``fonte_id`` (int) — restringe a uma fonte específica
- ``periodo`` (str) — ``"7d"``, ``"30d"``, ``"90d"``, ``"6m"``, ``"12m"``
  ou ``"15m"`` (Manual Cap. 4). Calcula ``data_inicio = hoje−N`` e filtra
  ``Verbatim.data_criacao_original >= data_inicio``. Vazio = tudo.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from flask import Blueprint, jsonify, request
from sqlalchemy import func

from src.api.engajamento import engajamento_escopo
from src.utils.sql import fmt_ano_mes
from src.auth import cliente_pode_ver_empresa
from src.models.local import Local
from src.models.verbatim import Verbatim
from src.utils.db import db_session


painel_bp = Blueprint("painel", __name__)


# ── Mapeamento subpilar → pilar ────────────────────────────────────────

PILAR_DE_SUBPILAR: Dict[str, str] = {
    "P1": "P",
    "P2": "P",
    "P3": "P",
    "D1": "D",
    "D2": "D",
    "D3": "D",
    "Pa1": "Pa",
    "Pa2": "Pa",
    "Pa3": "Pa",
    "A1": "A",
    "A2": "A",
    "A3": "A",
    # sem_lastro fica de fora dos 4 pilares (vai pra "outros")
}

PILARES_ORDEM = ["P", "D", "Pa", "A"]
# Nomenclatura oficial PDPA Loyall. Fonte canônica:
# data/PDPA_Manual_Operacao_v3.docx, Capítulo 2.
NOME_PILAR = {
    "P": "Precisão",
    "D": "Disponibilidade",
    "Pa": "Parceria",
    "A": "Aconselhamento",
}
SUBPILARES_ORDEM = [
    "P1",
    "P2",
    "P3",
    "D1",
    "D2",
    "D3",
    "Pa1",
    "Pa2",
    "Pa3",
    "A1",
    "A2",
    "A3",
]
# Manual PDPA v3, Capítulo 2 — nomes oficiais.
NOME_SUBPILAR = {
    "P1": "Calibração da Promessa",
    "P2": "Qualidade da Entrega",
    "P3": "Consistência ao Longo do Tempo",
    "D1": "Acessibilidade",
    "D2": "Eficácia Operacional",
    "D3": "Proatividade Estruturada",
    "Pa1": "Empatia Comercial",
    "Pa2": "Mutualidade",
    "Pa3": "Comprometimento Relacional",
    "A1": "Exemplo",
    "A2": "Orientação",
    "A3": "Recomendação Proativa",
}
TIPOS_ORDEM = ["promotor", "conversivel", "detrator", "inativo"]


# ── Ratio P/D (Manual Cap. 4) ─────────────────────────────────────────

RATIO_CAP_SUPERIOR = 9.99
RATIO_CAP_INFERIOR = 0.0


def calcular_ratio(promotor: int, detrator: int) -> float:
    """Ratio P/D conforme Manual Cap. 4.

    - Zero detratores → cap 9.99 (saturação positiva máxima).
    - Zero promotores → 0.0 (sinal crítico).
    - Caso normal: promotor / detrator, com cap em 9.99.
    """
    if promotor == 0 and detrator == 0:
        return 0.0
    if detrator == 0:
        return RATIO_CAP_SUPERIOR
    if promotor == 0:
        return RATIO_CAP_INFERIOR
    return min(RATIO_CAP_SUPERIOR, round(promotor / detrator, 2))


def _fmt_ratio_num(v: float) -> str:
    """1 casa decimal, vírgula pt-BR; inteiro quando exato (6.0→'6', 1.5→'1,5')."""
    r = round(v, 1)
    if r == int(r):
        return str(int(r))
    return f"{r:.1f}".replace(".", ",")


def ratio_em_palavras(ratio: float) -> str:
    """Tradução do ratio P/D em linguagem simples (CP-ratio-palavras).

    - ratio ≥ cap (9.99) → "sem detratores" (saturação positiva)
    - ratio ≤ 0          → "nenhum promotor"
    - ratio ≥ 1          → "X promotores para cada detrator"
    - ratio < 1          → "1 promotor para cada X detratores"
    X com 1 casa (vírgula pt-BR), inteiro quando exato; singular quando X=1.
    """
    if ratio >= RATIO_CAP_SUPERIOR:
        return "sem detratores"
    if ratio <= RATIO_CAP_INFERIOR:
        return "nenhum promotor"
    if ratio >= 1:
        promo = "promotor" if round(ratio, 1) == 1 else "promotores"
        return f"{_fmt_ratio_num(ratio)} {promo} para cada detrator"
    inv = 1 / ratio
    detr = "detrator" if round(inv, 1) == 1 else "detratores"
    return f"1 promotor para cada {_fmt_ratio_num(inv)} {detr}"


def _quarter_de(periodo: str):
    """``'YYYY-MM'`` → ``(ano, quarter 1..4)``. Q1=jan-mar … Q4=out-dez."""
    ano, mes = periodo.split("-")
    return int(ano), (int(mes) - 1) // 3 + 1


def historico_quarters_pilares(s, empresa_id, ag_id=None, local_id=None, n=4):
    """Histórico de ratio P/D por PILAR nos últimos ``n`` quarters COM dado, no
    escopo dado. Lê ``RatioMensal`` (série mensal por loja×subpilar) — $0, sem IA.

    - Ratio do quarter = ponderado por volume: ``Σ promotor ÷ Σ detrator`` dos 3
      meses, agregando os subpilares do pilar E as lojas do escopo (``calcular_ratio``).
    - Escopo: ``local_id`` (loja) tem precedência; senão ``ag_id`` (agrupamento);
      senão empresa inteira (agrega todas as lojas).
    - Retorna ``{pilar: [{"q": "Q3", "ratio": 0.28}, ...]}`` do mais antigo p/ o
      mais recente. Pilar com < 2 quarters é omitido (sem tendência)."""
    from collections import defaultdict

    from src.models.anomalia import RatioMensal

    q = s.query(
        RatioMensal.subpilar,
        RatioMensal.periodo,
        RatioMensal.promotor,
        RatioMensal.detrator,
        RatioMensal.total,
    ).filter(RatioMensal.empresa_id == empresa_id)
    if local_id is not None:
        q = q.filter(RatioMensal.local_id == local_id)
    elif ag_id is not None:
        q = q.filter(RatioMensal.agrupamento_id == ag_id)

    # (pilar, ano, quarter) -> [Σ promotor, Σ detrator, Σ total (todos os tipos)].
    # ``total`` é a coluna já existente de RatioMensal (mesmas linhas do ratio) —
    # nenhuma query/filtro novo, só uma soma a mais.
    acc: Dict[Any, List[int]] = defaultdict(lambda: [0, 0, 0])
    for sub, periodo, prom, det, tot in q.all():
        pilar = PILAR_DE_SUBPILAR.get(sub)
        if pilar is None or not periodo:
            continue
        ano, quarter = _quarter_de(periodo)
        chave = (pilar, ano, quarter)
        acc[chave][0] += prom or 0
        acc[chave][1] += det or 0
        acc[chave][2] += tot or 0

    por_pilar: Dict[str, List] = defaultdict(list)
    for (pilar, ano, quarter), (prom, det, tot) in acc.items():
        por_pilar[pilar].append((ano, quarter, prom, det, tot))

    out: Dict[str, List[Dict[str, Any]]] = {}
    for pilar, linhas in por_pilar.items():
        linhas.sort()  # (ano, quarter) crescente → mais antigo primeiro
        ultimos = linhas[-n:]
        if len(ultimos) < 2:
            continue  # sem tendência
        out[pilar] = [
            {
                "q": f"Q{quarter}",
                "ano": ano,
                "chave": f"{ano}Q{quarter}",  # p/ o hx-get do drawer de detalhe
                "ratio": calcular_ratio(prom, det),
                "total": tot,  # N de verbatins do quarter (todos os tipos)
            }
            for (ano, quarter, prom, det, tot) in ultimos
        ]
    return out


def _subpilares_do_pilar(pilar: str) -> List[str]:
    return [sp for sp in SUBPILARES_ORDEM if PILAR_DE_SUBPILAR.get(sp) == pilar]


def _meses_do_quarter(ano: int, quarter: int) -> List[str]:
    """['YYYY-MM', x3] dos meses do quarter (Q1=jan-mar … Q4=out-dez)."""
    m0 = (quarter - 1) * 3 + 1
    return [f"{ano}-{m:02d}" for m in (m0, m0 + 1, m0 + 2)]


def _intervalo_quarter(ano: int, quarter: int):
    """(date_inicio, date_fim_exclusivo) do quarter — p/ filtrar datetime."""
    from datetime import date

    m0 = (quarter - 1) * 3 + 1
    ini = date(ano, m0, 1)
    fim = date(ano + 1, 1, 1) if quarter == 4 else date(ano, m0 + 3, 1)
    return ini, fim


def _serie_quarters_pilar(s, empresa_id, pilar, ag_id, local_id):
    """[(ano, quarter, Σprom, Σdet)] do pilar no escopo, ordenado crescente."""
    from collections import defaultdict

    from src.models.anomalia import RatioMensal

    subs = _subpilares_do_pilar(pilar)
    q = s.query(RatioMensal.periodo, RatioMensal.promotor, RatioMensal.detrator).filter(
        RatioMensal.empresa_id == empresa_id, RatioMensal.subpilar.in_(subs)
    )
    if local_id is not None:
        q = q.filter(RatioMensal.local_id == local_id)
    elif ag_id is not None:
        q = q.filter(RatioMensal.agrupamento_id == ag_id)
    acc: Dict[Any, List[int]] = defaultdict(lambda: [0, 0])
    for periodo, prom, det in q.all():
        if not periodo:
            continue
        ano, quarter = _quarter_de(periodo)
        acc[(ano, quarter)][0] += prom or 0
        acc[(ano, quarter)][1] += det or 0
    return sorted((ano, qu, pd[0], pd[1]) for (ano, qu), pd in acc.items())


def _loja_mais_impactou(s, empresa_id, pilar, sel, prev, ag_id, local_id):
    """Loja de maior contribuição (volume × variação de ratio) à variação do pilar
    entre o quarter ``prev`` e ``sel`` (cada um ``(ano, quarter)``). Só lojas com
    dado nos dois quarters. Retorna {nome, ratio, variacao} ou None."""
    from collections import defaultdict

    from src.models.anomalia import RatioMensal
    from src.models.local import Local

    subs = _subpilares_do_pilar(pilar)
    meses_sel = set(_meses_do_quarter(*sel))
    meses_prev = set(_meses_do_quarter(*prev))
    q = s.query(
        RatioMensal.local_id,
        RatioMensal.periodo,
        RatioMensal.promotor,
        RatioMensal.detrator,
    ).filter(
        RatioMensal.empresa_id == empresa_id,
        RatioMensal.subpilar.in_(subs),
        RatioMensal.local_id.isnot(None),
        RatioMensal.periodo.in_(meses_sel | meses_prev),
    )
    if local_id is not None:
        q = q.filter(RatioMensal.local_id == local_id)
    elif ag_id is not None:
        q = q.filter(RatioMensal.agrupamento_id == ag_id)

    sel_acc: Dict[int, List[int]] = defaultdict(lambda: [0, 0])
    prev_acc: Dict[int, List[int]] = defaultdict(lambda: [0, 0])
    for lid, periodo, prom, det in q.all():
        alvo = sel_acc if periodo in meses_sel else prev_acc
        alvo[lid][0] += prom or 0
        alvo[lid][1] += det or 0

    melhor = None  # (|contrib|, lid, ratio_sel, variacao)
    for lid, (prom, det) in sel_acc.items():
        if lid not in prev_acc:
            continue  # precisa dos 2 quarters p/ ter variação
        r_sel = calcular_ratio(prom, det)
        r_prev = calcular_ratio(prev_acc[lid][0], prev_acc[lid][1])
        contrib = (prom + det) * (r_sel - r_prev)  # volume × variação
        if melhor is None or abs(contrib) > melhor[0]:
            melhor = (abs(contrib), lid, r_sel, round(r_sel - r_prev, 2))
    if melhor is None:
        return None
    _, lid, r_sel, var = melhor
    loc = s.get(Local, lid)
    return {"nome": loc.nome if loc else f"loja {lid}", "ratio": r_sel, "variacao": var}


def _tema_dominante_periodo(s, empresa_id, pilar, ano, quarter, ag_id, local_id):
    """Tema mais frequente nos detratores do pilar no quarter (Verbatim×VerbatimTema
    ×Tema por data). Retorna {nome, volume} ou None."""
    from sqlalchemy import func as _func

    from src.models.local import Local
    from src.models.temas import Tema, VerbatimTema
    from src.models.verbatim import Verbatim

    subs = _subpilares_do_pilar(pilar)
    ini, fim = _intervalo_quarter(ano, quarter)
    q = (
        s.query(Tema.nome, _func.count(Verbatim.id))
        .join(VerbatimTema, VerbatimTema.tema_id == Tema.id)
        .join(Verbatim, Verbatim.id == VerbatimTema.verbatim_id)
        .filter(
            Verbatim.empresa_id == empresa_id,
            Verbatim.subpilar.in_(subs),
            Verbatim.tipo == "detrator",
            Verbatim.data_criacao_original >= ini,
            Verbatim.data_criacao_original < fim,
        )
    )
    if local_id is not None:
        q = q.filter(Verbatim.local_id == local_id)
    elif ag_id is not None:
        q = q.filter(Verbatim.local_id.in_(s.query(Local.id).filter(Local.agrupamento_id == ag_id)))
    linha = q.group_by(Tema.nome).order_by(_func.count(Verbatim.id).desc()).first()
    return {"nome": linha[0], "volume": int(linha[1])} if linha else None


def _anomalia_periodo(s, empresa_id, pilar, ano, quarter, ag_id, local_id):
    """AnomaliaDetectada (severidade ≥ atenção) do pilar no quarter. crítico antes
    de atenção. Retorna {titulo, severidade} ou None."""
    from src.models.anomalia import AnomaliaDetectada
    from src.models.local import Local

    subs = _subpilares_do_pilar(pilar)
    q = s.query(AnomaliaDetectada).filter(
        AnomaliaDetectada.empresa_id == empresa_id,
        AnomaliaDetectada.subpilar.in_(subs),
        AnomaliaDetectada.periodo.in_(_meses_do_quarter(ano, quarter)),
        AnomaliaDetectada.severidade.in_(("atencao", "critico")),
    )
    if local_id is not None:
        q = q.filter(AnomaliaDetectada.local_id == local_id)
    elif ag_id is not None:
        q = q.filter(AnomaliaDetectada.agrupamento_id == ag_id)
    rows = q.all()
    if not rows:
        return None
    ordem = {"critico": 0, "atencao": 1}
    a = min(rows, key=lambda x: (ordem.get(x.severidade, 9), x.periodo or ""))
    nome_sub = NOME_SUBPILAR.get(a.subpilar, a.subpilar)
    loc = s.get(Local, a.local_id) if a.local_id else None
    titulo = f"{loc.nome} · {nome_sub}" if loc else nome_sub
    return {"titulo": titulo, "severidade": a.severidade}


def quarter_detalhe_pilar(s, empresa_id, pilar, ano, quarter, ag_id=None, local_id=None):
    """Detalhe de um quarter de um pilar p/ o drawer (CP-quarter-detalhe), $0:
    variação vs quarter anterior, loja que mais impactou, tema dominante e anomalia.
    Retorna ``None`` se o quarter não tem dado no escopo."""
    serie = _serie_quarters_pilar(s, empresa_id, pilar, ag_id, local_id)
    idx = next((i for i, (a, qu, _, _) in enumerate(serie) if a == ano and qu == quarter), None)
    if idx is None:
        return None
    _a, _qu, prom, det = serie[idx]
    ratio = calcular_ratio(prom, det)

    variacao = None
    loja = None
    if idx > 0:
        pa, pqu, pprom, pdet = serie[idx - 1]
        variacao = {
            "delta": round(ratio - calcular_ratio(pprom, pdet), 2),
            "quarter_anterior": f"Q{pqu}",
        }
        loja = _loja_mais_impactou(s, empresa_id, pilar, (ano, quarter), (pa, pqu), ag_id, local_id)

    return {
        "pilar": pilar,
        "pilar_nome": NOME_PILAR.get(pilar, pilar),
        "quarter_label": f"Q{quarter}",
        "ano": ano,
        "ratio": ratio,
        "variacao": variacao,
        "loja": loja,
        "tema": _tema_dominante_periodo(s, empresa_id, pilar, ano, quarter, ag_id, local_id),
        "anomalia": _anomalia_periodo(s, empresa_id, pilar, ano, quarter, ag_id, local_id),
    }


# Faixas operacionais do ratio — verdade única (ver docs/PROJETO_PDPA.md).
# Lista ordenada de (limite_superior_exclusivo, label); o último (inf) é o teto.
# Centralizado no CP-LG-0 para reuso pela Lente de Governança. NÃO alterar os
# cortes sem reconciliar PROJETO_PDPA.md + a escala Proximity (que é separada).
FAIXAS_RATIO = (
    (0.5, "critico"),
    (1.0, "fraco"),
    (2.0, "atencao"),
    (5.0, "bom"),
    (float("inf"), "excelente"),
)


def faixa_ratio(ratio: float) -> str:
    """Devolve a faixa semântica do ratio (5 níveis, cores do painel).

    - 0.0–0.5  : critico       (vermelho)
    - 0.5–1.0  : fraco         (laranja)
    - 1.0–2.0  : atencao       (amarelo)
    - 2.0–5.0  : bom           (verde claro)
    - ≥ 5.0    : excelente     (verde escuro)
    """
    for limite, label in FAIXAS_RATIO:
        if ratio < limite:
            return label
    return FAIXAS_RATIO[-1][1]  # inalcançável: o último limite é inf


def ratios_por_pilar(agg: Dict[str, Dict[str, Any]]) -> Dict[str, float]:
    """Ratio P/D agregado por pilar, só para os pilares COM volume (prom+det>0).

    Volume = prom+det (não ``total``): pilar só com conversíveis/inativos não tem
    sinal P/D — ``calcular_ratio(0,0)`` daria 0.0 e o faria falso-crítico. ``agg`` =
    ``{subpilar: {prom, det, …}}`` (saída de ``agregar_subpilares``)."""
    por: Dict[str, Dict[str, int]] = {}
    for sub, d in agg.items():
        p = PILAR_DE_SUBPILAR.get(sub)
        if not p:
            continue
        x = por.setdefault(p, {"prom": 0, "det": 0})
        x["prom"] += d["prom"]
        x["det"] += d["det"]
    return {
        p: calcular_ratio(x["prom"], x["det"]) for p, x in por.items() if (x["prom"] + x["det"]) > 0
    }


def gargalo_sequencial(agg: Dict[str, Dict[str, Any]]) -> Optional[str]:
    """Gargalo do Lastro — regra SEQUENCIAL em duas camadas (definição de método).

    A jornada P→D→Pa→A trava no primeiro elo QUEBRADO. O crítico tem PRECEDÊNCIA
    sobre a posição (é a quebra que trava tudo); só na ausência de crítico o primeiro
    fraco assume. Dentro de cada camada, o primeiro na ordem canônica:

    1. primeiro pilar (com volume) CRÍTICO — ``ratio < 0.5``;
    2. senão, primeiro pilar FRACO — ``0.5 ≤ ratio < 1.0``;
    3. senão (nada abaixo de 1.0), ``None`` — nada quebrado, não há gargalo.

    Limiares vêm de ``FAIXAS_RATIO`` (fonte única). Substitui a regra ANTIGA de
    "menor ratio" (posicional-agnóstica), que contradizia o cabeçalho sequencial do
    Lastro (ex.: Aconselhamento, último pilar, virava gargalo por ter o menor ratio)."""
    ratios = ratios_por_pilar(agg)
    criticos = [p for p in PILARES_ORDEM if p in ratios and ratios[p] < 0.5]
    if criticos:
        return criticos[0]
    fracos = [p for p in PILARES_ORDEM if p in ratios and ratios[p] < 1.0]
    return fracos[0] if fracos else None


# ── Métricas consolidadas (Manual Cap. 4) ─────────────────────────────


def _normalizar_indice(base: float) -> float:
    """Normaliza o ratio-base (0–9.99) para a escala 0-10 do Índice Geral, ANCORADA na
    régua de ratio do Manual v8 — por partes, não linear: ratio 1,0 (empate) → 5 (piso
    'atenção'), 2,0 ('bom') → 7 (piso 'saudável'), 5,0 ('excelente') → 10. Assim as faixas
    do índice (≥7/5-7/<5) concordam com a régua de ratio dos pilares. Substitui o ×2 do
    hotfix, que punha ratio 'bom' (2,0) em índice 4,0 = 'crítico' (choque com a régua)."""
    if base <= 0.5:
        return round(base * 5.0, 2)  # 0,5 → 2,5
    if base <= 1.0:
        return round(2.5 + (base - 0.5) * 5.0, 2)  # 1,0 → 5,0 (piso atenção = empate)
    if base <= 2.0:
        return round(5.0 + (base - 1.0) * 2.0, 2)  # 2,0 → 7,0 (piso saudável = 'bom')
    if base <= 5.0:
        return round(7.0 + (base - 2.0), 2)  # 5,0 → 10 (excelente)
    return 10.0


def _base_indice(matriz_subpilares, pilares=None):
    """``(base, ratio_pior_pilar, ratio_medio_ponderado)`` — fonte única do cálculo.
    ``base = min(pior, média)`` (elo mais fraco). ``pior`` = None quando não há pilar
    mensurável (sem volume) → base 0.0. Usado por ``calcular_indice_geral`` e pelo flag
    ``indice_governado_pelo_pior`` (sem duplicar a extração)."""
    total_volume = sum(c.get("total", 0) for c in matriz_subpilares)
    if total_volume == 0:
        return (0.0, None, 0.0)
    soma = sum(c.get("ratio", 0.0) * c.get("total", 0) for c in matriz_subpilares)
    media = soma / total_volume
    # Ratio do pior pilar — só pilares com volume > 0.
    if pilares is not None:
        ratios = [p["ratio"] for p in pilares if p.get("total", 0) > 0]
    else:  # agrega da matriz por prefixo do código (P/D/Pa/A)
        agg: Dict[str, Dict[str, int]] = {}
        for c in matriz_subpilares:
            p_code = PILAR_DE_SUBPILAR.get(c.get("subpilar", ""))
            if p_code is None:
                continue
            d = agg.setdefault(p_code, {"promotor": 0, "detrator": 0, "total": 0})
            d["promotor"] += c.get("promotor", 0)
            d["detrator"] += c.get("detrator", 0)
            d["total"] += c.get("total", 0)
        ratios = [
            calcular_ratio(d["promotor"], d["detrator"]) for d in agg.values() if d["total"] > 0
        ]
    if not ratios:
        return (0.0, None, media)
    pior = min(ratios)
    return (min(pior, media), pior, media)


def calcular_indice_geral(
    matriz_subpilares: List[Dict[str, Any]],
    pilares: Optional[List[Dict[str, Any]]] = None,
) -> float:
    """Índice Geral (escala 0-10) — Manual v8 (Índice Geral).

    Fórmula:
        base   = min(ratio_pior_pilar, ratio_medio_ponderado)   # elo mais fraco
        indice = _normalizar_indice(base)                        # régua de ratio → 0-10

    O ``min`` (elo mais fraco) impede que um pilar saturado (Pa1 9.99, alto volume)
    mascare um pilar crítico via média — masking que é a MAIORIA da base, não caso raro
    do BH Airport. A normalização é ancorada na régua de ratio (1,0→5, 2,0→7, 5,0→10)
    para as faixas do índice (≥7/5-7/<5) concordarem com a régua dos pilares: ratio 'bom'
    (2,0) → 'saudável', não 'crítico' (o ×2 antigo o punha em índice 4,0). Manual v8.

    Args:
        matriz_subpilares: lista de dicts {ratio, total} por subpilar
            (12 entradas). Usado pra calcular o ratio médio ponderado.
        pilares: lista de 4 dicts {pilar, ratio, total} por pilar
            (P/D/Pa/A). Usado pra extrair o ratio do pior pilar.
            Quando None, calcula da matriz agregando por prefixo.

    Returns:
        0.0 quando não há volume.
    """
    base, _, _ = _base_indice(matriz_subpilares, pilares)
    return _normalizar_indice(base)


def indice_governado_pelo_pior(matriz_subpilares, pilares=None) -> bool:
    """True quando o PIOR PILAR (não a média) define o índice — o ``min`` pegou o pior
    pilar (pior ≤ média) e há pilar mensurável. Alimenta a nota do card ('Aqui, {pilar}
    é o teto') sem o template recomputar o índice (que agora é normalização por partes,
    não ``ratio×2``)."""
    _, pior, media = _base_indice(matriz_subpilares, pilares)
    return pior is not None and pior <= media


def faixa_indice_geral(indice: float) -> str:
    """Faixa do Índice Geral (Manual Cap. 4): ≥7 saudavel, 5-7 atencao, <5 critico."""
    if indice >= 7.0:
        return "saudavel"
    if indice >= 5.0:
        return "atencao"
    return "critico"


def indice_pdpa(pilares, codigos=None):
    """Índice PDPA (0-100) + volume classificado do recorte — a relação em um número.

    ``(promotores + conversíveis·0,5) / (prom+conv+det) · 100``. O conversível conta
    metade (relação incompleta, não ausência); o detrator fica no denominador. O
    denominador é ``prom+conv+det`` explícito → inativo e sem_lastro ficam fora
    (sem_lastro não entra em nenhum dos 4 pilares). ``codigos=None`` → Índice PDPA
    (todos); ``{"P","D"}`` → Base; ``{"Pa","A"}`` → Topo. Retorna ``(None, 0)`` sem
    volume classificado no recorte. Manual Cap. 4 (Índice PDPA)."""
    sel = [p for p in pilares if codigos is None or p["pilar"] in codigos]
    num = sum(p["promotor"] + p["conversivel"] * 0.5 for p in sel)
    den = sum(p["promotor"] + p["conversivel"] + p["detrator"] for p in sel)
    return (round(num / den * 100, 1), int(den)) if den else (None, 0)


# Horizonte de LEITURA da Previsibilidade: os N meses mais recentes. Ela mede a
# consistência da operação ATUAL — série de década lia deriva secular como
# instabilidade (Club Med Brasil: 42→68 ao passar de 115 meses para 12). 12 é o
# mínimo defensável: >=1 ciclo sazonal, e o piso de 3 meses deixa margem. Não é o
# número da COLETA (15m) nem da RETENÇÃO (18m) — é horizonte de leitura próprio.
# Ordenação por mês DESC, pega os N mais recentes: determinístico (senão o descarte
# seria arbitrário e a nota oscilaria entre execuções sobre o MESMO dado).
JANELA_PREVISIBILIDADE_MESES = 12


def calcular_previsibilidade(
    empresa_id: int,
    s,
    base_query_args: Dict[str, Any],
) -> Optional[float]:
    """Previsibilidade (escala 0-100) conforme Manual Cap. 4: ``1 − CV`` dos ratios,
    homogeneidade entre lojas e no tempo.

    Realinhada 2026-08-08 (dispersão pura): removidos 3 desvios não-Manual que
    vieram da cópia atacadão do v2 no hotfix 99011d4 — (1) o termo
    ``pct_conversíveis·0,3`` (aproveitamento de dado, não dispersão), (2) o fator
    ``/2`` em ``min(CV/2,1)`` (escala inventada, o ``×2`` do Índice de novo), (3) o
    eixo não-medido virava "1" fantasma. Fórmula agora:

        eixo_lojas = 1 − min(CV(ratios_locais), 1)   # >= 2 lojas c/ >= 5 verb
        eixo_tempo = 1 − min(CV(ratios_meses), 1)    # >= 3 meses c/ >= 3 verb
        score = média dos eixos COM BASE × 100       # eixo sem base sai da conta

    Sem nenhum eixo com base → ``None`` (não há default). Filtros do painel
    aplicados (agrupamento, local, fonte, período).

    ``sem_lastro`` (=inativo) fica FORA dos dois eixos (2026-08-10): já não entrava
    no ratio (prom/det), mas inflava o piso (>=5 lojas / >=3 meses), deixando lojas
    de puro ruído qualificarem a dispersão. Efeito depende de como o ruído se
    distribui, não de quanto existe — concentrado (poucas lojas) distorce; espalhado
    não. NULL (pendente) segue contando o piso.

    Janela de leitura (2026-08-11): os **12 meses mais recentes** nos DOIS eixos
    (``JANELA_PREVISIBILIDADE_MESES``). Mede a operação atual, não a década — série longa
    lia deriva secular como instabilidade. Corte determinístico (mês DESC, N mais recentes).
    """
    import statistics
    from datetime import datetime

    # 0. Janela de 12 meses (horizonte de leitura). Determina os N meses mais recentes
    #    COM dado (ordem DESC), âncora no próprio dado (imune à pausa de coleta), e corta
    #    AMBOS os eixos no início do mais antigo deles — média de horizontes diferentes
    #    não significaria nada. NULL sem data já não entra no eixo tempo.
    mes_expr = fmt_ano_mes(Verbatim.data_criacao_original)
    q_meses_distintos = (
        s.query(mes_expr)
        .filter(Verbatim.empresa_id == empresa_id)
        .filter(Verbatim.data_criacao_original.isnot(None))
        .filter(func.coalesce(Verbatim.subpilar, "") != "sem_lastro")
        .group_by(mes_expr)
    )
    meses_recentes = sorted(
        (
            m
            for (m,) in _apply_query_args(q_meses_distintos, empresa_id, s, base_query_args).all()
            if m
        ),
        reverse=True,
    )[:JANELA_PREVISIBILIDADE_MESES]
    corte_janela = None
    if meses_recentes:
        _ano, _mes = min(meses_recentes).split("-")
        corte_janela = datetime(int(_ano), int(_mes), 1)

    # 1. Ratios por local (lojas) — usa só locais com >= 5 verbatins
    #    sem_lastro (=inativo) NÃO entra: já está fora do ratio (prom/det), e deixá-lo
    #    inflar o piso >= 5 faz lojas de puro ruído qualificarem a dispersão. Regra
    #    "sem_lastro não entra em conta". NULL (pendente de classificação) fica.
    q_locais = (
        s.query(
            Verbatim.local_id,
            Verbatim.tipo,
            func.count(Verbatim.id),
        )
        .filter(Verbatim.empresa_id == empresa_id)
        .filter(Verbatim.local_id.isnot(None))
        .filter(func.coalesce(Verbatim.subpilar, "") != "sem_lastro")
        .group_by(Verbatim.local_id, Verbatim.tipo)
    )
    if corte_janela is not None:
        q_locais = q_locais.filter(Verbatim.data_criacao_original >= corte_janela)
    rows_locais = _apply_query_args(q_locais, empresa_id, s, base_query_args).all()
    por_local: Dict[int, Dict[str, int]] = {}
    for lid, tipo, qtd in rows_locais:
        d = por_local.setdefault(lid, {"promotor": 0, "detrator": 0, "total": 0})
        d["total"] += qtd
        if tipo in d:
            d[tipo] += qtd
    ratios_locais = [
        calcular_ratio(d["promotor"], d["detrator"]) for d in por_local.values() if d["total"] >= 5
    ]

    # 2. Ratios por mês — usa só meses com >= 3 verbatins (sem_lastro fora do piso, idem eixo 1);
    #    janelado nos 12 mais recentes (mesmo corte do eixo lojas).
    q_meses = (
        s.query(mes_expr.label("mes"), Verbatim.tipo, func.count(Verbatim.id))
        .filter(Verbatim.empresa_id == empresa_id)
        .filter(Verbatim.data_criacao_original.isnot(None))
        .filter(func.coalesce(Verbatim.subpilar, "") != "sem_lastro")
        .group_by(mes_expr, Verbatim.tipo)
    )
    if corte_janela is not None:
        q_meses = q_meses.filter(Verbatim.data_criacao_original >= corte_janela)
    rows_meses = _apply_query_args(q_meses, empresa_id, s, base_query_args).all()
    por_mes: Dict[str, Dict[str, int]] = {}
    for mes, tipo, qtd in rows_meses:
        d = por_mes.setdefault(mes, {"promotor": 0, "detrator": 0, "total": 0})
        d["total"] += qtd
        if tipo in d:
            d[tipo] += qtd
    ratios_meses = [
        calcular_ratio(d["promotor"], d["detrator"]) for d in por_mes.values() if d["total"] >= 3
    ]

    # 3. Um eixo por dispersão medível: 1 − min(CV, 1). Eixo sem base fica de fora
    # (não vira 1 fantasma) — renormalização automática pela média dos eixos com base.
    eixos = []
    if len(ratios_locais) >= 2:
        cv = statistics.stdev(ratios_locais) / max(statistics.mean(ratios_locais), 0.01)
        eixos.append(1.0 - min(cv, 1.0))
    if len(ratios_meses) >= 3:
        cv = statistics.stdev(ratios_meses) / max(statistics.mean(ratios_meses), 0.01)
        eixos.append(1.0 - min(cv, 1.0))

    # 4. Média dos eixos com base; sem nenhum eixo medível → None (não há default).
    if not eixos:
        return None
    score = (sum(eixos) / len(eixos)) * 100
    return round(max(0.0, min(100.0, score)), 1)


def _apply_query_args(q, empresa_id: int, s, base_query_args: Dict[str, Any]):
    """Reaplica os filtros painel numa query base (uso interno de
    calcular_previsibilidade e calcular_concentracao_detratores)."""
    if base_query_args.get("agrupamento_id"):
        try:
            ag_id = int(base_query_args["agrupamento_id"])
            locais_do_ag = [
                lid
                for (lid,) in s.query(Local.id)
                .filter_by(empresa_id=empresa_id, agrupamento_id=ag_id)
                .all()
            ]
            q = q.filter(Verbatim.local_id.in_(locais_do_ag or [-1]))
        except (ValueError, TypeError):
            pass
    if base_query_args.get("local_id"):
        try:
            q = q.filter(Verbatim.local_id == int(base_query_args["local_id"]))
        except (ValueError, TypeError):
            pass
    if base_query_args.get("fonte_id"):
        try:
            q = q.filter(Verbatim.fonte_id == int(base_query_args["fonte_id"]))
        except (ValueError, TypeError):
            pass
    if base_query_args.get("data_inicio_periodo"):
        q = q.filter(Verbatim.data_criacao_original >= base_query_args["data_inicio_periodo"])
    return q


# Horizonte de LEITURA da Concentração e do Gini: os N meses mais recentes. Ambos
# respondem "onde intervir AGORA" (decisão de alocação recorrente), não posição
# estrutural — all-time diluía a dor recente (Localiza: 54%→64% top-5 ao janelar; a
# dor recente é MAIS concentrada). 6m = mesma família dos temas ("o que está vivo
# agora"). Cross-sectional (compara lojas entre si), não precisa de série longa. Corte
# = MAX(data_criacao)−6m por empresa (imune à pausa de coleta). Ver [[project_horizontes]].
JANELA_CONCENTRACAO_MESES = 6


def _corte_janela_meses(empresa_id: int, s, meses: int) -> Optional[datetime]:
    """Início da janela de ``meses`` ancorada no dado mais recente da empresa:
    ``MAX(data_criacao_original) − meses`` (não em "hoje" — imune à pausa de coleta).
    ``None`` se a empresa não tem dado datado (→ sem corte)."""
    maxd = (
        s.query(func.max(Verbatim.data_criacao_original))
        .filter(Verbatim.empresa_id == empresa_id)
        .scalar()
    )
    return (maxd - timedelta(days=meses * 30)) if maxd is not None else None


def calcular_concentracao_detratores(
    empresa_id: int, s, base_query_args: Dict[str, Any]
) -> Optional[float]:
    """Concentração de Detratores (%) conforme Manual Cap. 4.

    Ranqueia locais ascendentemente por ratio (piores primeiro). Soma os
    detratores das 5 piores lojas e divide pelo total de detratores da
    empresa, em %.

    Devolve ``None`` se a empresa não tem locais suficientes para
    interpretação (>0 mas <5 locais com volume — métrica perde sentido).

    > 60% = cirúrgico (poucas lojas concentram o problema).
    < 30% = sistêmico (distribuído, processo central).
    """
    # Agrega por local: total de promotores e detratores em verbatins
    # da empresa (com filtros do painel ja aplicados via base_query).
    q = (
        s.query(
            Verbatim.local_id,
            Verbatim.tipo,
            func.count(Verbatim.id),
        )
        .filter(Verbatim.empresa_id == empresa_id)
        .filter(Verbatim.local_id.isnot(None))
        .group_by(Verbatim.local_id, Verbatim.tipo)
    )

    # Aplica filtros opcionais que vieram do request (mesma assinatura
    # de _aplicar_filtros, mas tem que ser inline aqui pra não duplicar)
    if base_query_args.get("agrupamento_id"):
        try:
            ag_id = int(base_query_args["agrupamento_id"])
            locais_do_ag = [
                lid
                for (lid,) in s.query(Local.id)
                .filter_by(empresa_id=empresa_id, agrupamento_id=ag_id)
                .all()
            ]
            q = q.filter(Verbatim.local_id.in_(locais_do_ag or [-1]))
        except (ValueError, TypeError):
            pass
    if base_query_args.get("local_id"):
        try:
            q = q.filter(Verbatim.local_id == int(base_query_args["local_id"]))
        except (ValueError, TypeError):
            pass
    if base_query_args.get("fonte_id"):
        try:
            q = q.filter(Verbatim.fonte_id == int(base_query_args["fonte_id"]))
        except (ValueError, TypeError):
            pass
    # Janela de leitura: período explícito do request vence; senão, default de 6 meses
    # ancorado no dado (MAX(data)−6m, imune à pausa). A Concentração passa a medir a dor
    # ATUAL, não a década — all-time diluía (ver JANELA_CONCENTRACAO_MESES).
    if base_query_args.get("data_inicio_periodo"):
        q = q.filter(Verbatim.data_criacao_original >= base_query_args["data_inicio_periodo"])
    else:
        corte = _corte_janela_meses(empresa_id, s, JANELA_CONCENTRACAO_MESES)
        if corte is not None:
            q = q.filter(Verbatim.data_criacao_original >= corte)

    rows = q.all()

    # Constrói (local_id → {promotor, detrator, total})
    por_local: Dict[int, Dict[str, int]] = {}
    for local_id, tipo, qtd in rows:
        d = por_local.setdefault(local_id, {"promotor": 0, "detrator": 0, "total": 0})
        d["total"] += qtd
        if tipo in d:
            d[tipo] += qtd

    # Filtro de volume mínimo (Manual: "ratio confiável precisa volume").
    # v2 usava >= 5; mesmo critério.
    LOCAL_VOLUME_MIN = 5
    locais_com_volume = [(lid, d) for lid, d in por_local.items() if d["total"] >= LOCAL_VOLUME_MIN]
    total_locais = len(locais_com_volume)
    total_detratores = sum(d["detrator"] for _, d in locais_com_volume)
    if total_locais < 5 or total_detratores == 0:
        return None

    # Ranqueia ascendentemente por ratio (piores primeiro)
    def _ratio(d: Dict[str, int]) -> float:
        return calcular_ratio(d["promotor"], d["detrator"])

    locais_com_volume.sort(key=lambda x: _ratio(x[1]))
    piores_5 = locais_com_volume[:5]
    detratores_top5 = sum(d["detrator"] for _, d in piores_5)
    return round(100.0 * detratores_top5 / total_detratores, 1)


def faixa_concentracao(pct: Optional[float]) -> str:
    """Faixa da Concentração de Detratores (Manual Cap. 4).

    - > 60%: cirurgico (intervenção em poucas lojas resolve)
    - 30-60%: misto
    - < 30%: sistemico (processo central precisa revisão)
    - None: indisponivel (< 5 locais com volume, ou zero detratores)
    """
    if pct is None:
        return "indisponivel"
    if pct > 60.0:
        return "cirurgico"
    if pct >= 30.0:
        return "misto"
    return "sistemico"


# ── Guards de LEITURA dos cards (T1/T2) — queries à parte, cálculo INTOCADO ──────
CONCENTRACAO_MIN_LOJAS_LEITURA = 10  # abaixo disto o top-5 é a maioria das lojas → trivial


def previsibilidade_medida(empresa_id: int, s, base_query_args: Dict[str, Any]) -> bool:
    """Há dispersão MEDÍVEL em ≥1 eixo? (≥2 lojas com ≥5 verbatins OU ≥3 meses com ≥3 —
    as MESMAS condições de calcular_previsibilidade:616,621). Distingue previsibilidade
    medida do default 70,0 (var_locais=var_temporal=0). NÃO toca o cálculo."""
    q_loc = (
        s.query(Verbatim.local_id, func.count(Verbatim.id))
        .filter(Verbatim.empresa_id == empresa_id, Verbatim.local_id.isnot(None))
        .group_by(Verbatim.local_id)
    )
    n_lojas = sum(
        1 for _, c in _apply_query_args(q_loc, empresa_id, s, base_query_args).all() if c >= 5
    )
    if n_lojas >= 2:
        return True
    mes_expr = fmt_ano_mes(Verbatim.data_criacao_original)
    q_mes = (
        s.query(mes_expr, func.count(Verbatim.id))
        .filter(Verbatim.empresa_id == empresa_id, Verbatim.data_criacao_original.isnot(None))
        .group_by(mes_expr)
    )
    n_meses = sum(
        1 for _, c in _apply_query_args(q_mes, empresa_id, s, base_query_args).all() if c >= 3
    )
    return n_meses >= 3


def concentracao_n_lojas(empresa_id: int, s, base_query_args: Dict[str, Any]) -> int:
    """Nº de lojas com ≥5 verbatins (mesma base/piso de calcular_concentracao_detratores:730).
    O cálculo só devolve o pct; esta contagem alimenta o guard T2 (top-5 trivial com poucas
    lojas). NÃO toca o cálculo."""
    q = (
        s.query(Verbatim.local_id, func.count(Verbatim.id))
        .filter(Verbatim.empresa_id == empresa_id, Verbatim.local_id.isnot(None))
        .group_by(Verbatim.local_id)
    )
    return sum(1 for _, c in _apply_query_args(q, empresa_id, s, base_query_args).all() if c >= 5)


# ── Texto descritivo do escopo (Bloco 5 hotfix UI) ────────────────────

# Mapa amigável de conector → nome legível para a descrição do escopo.
NOME_FONTE_AMIGAVEL = {
    "google": "Google Reviews",
    "google_news": "Google News",
    "tripadvisor": "TripAdvisor",
    "instagram": "Instagram",
    "facebook": "Facebook",
    "linkedin": "LinkedIn",
    "youtube": "YouTube",
    "tiktok": "TikTok",
    "appstore": "App Store / Play Store",
    "mercadolivre": "Mercado Livre",
    "website": "website",
    "glassdoor": "Glassdoor",
    "indeed": "Indeed",
    "excel_manual": "Excel manual",
}

# Labels amigáveis para período (mesmas chaves do _resolver_periodo).
PERIODO_LABEL = {
    "7d": "nos últimos 7 dias",
    "30d": "nos últimos 30 dias",
    "90d": "nos últimos 90 dias",
    "6m": "nos últimos 6 meses",
    "12m": "nos últimos 12 meses",
    "15m": "nos últimos 15 meses",
}


def descrever_escopo(
    empresa_nome: str,
    *,
    agrupamento_nome: Optional[str] = None,
    local_nome: Optional[str] = None,
    fonte_conector: Optional[str] = None,
    periodo: Optional[str] = None,
) -> str:
    """Compõe o sufixo descritivo dos filtros aplicados no painel.

    Regras:
    - Espacial: ``local`` sobrescreve ``agrupamento``; sem nada → "geral
      da {empresa_nome}".
    - Fonte: acrescenta "via {nome amigável}".
    - Período: acrescenta "nos últimos N".

    Exemplos::

        descrever_escopo("BH Airport")
        # → "geral da BH Airport"

        descrever_escopo("BH Airport", agrupamento_nome="Aeroporto")
        # → "no agrupamento Aeroporto"

        descrever_escopo("BH Airport", local_nome="Terminal Confins",
                         fonte_conector="google", periodo="7d")
        # → "em Terminal Confins via Google Reviews nos últimos 7 dias"
    """
    partes = []
    if local_nome:
        partes.append(f"em {local_nome}")
    elif agrupamento_nome:
        partes.append(f"no agrupamento {agrupamento_nome}")
    else:
        partes.append(f"geral da {empresa_nome}")

    if fonte_conector:
        nome = NOME_FONTE_AMIGAVEL.get(fonte_conector, fonte_conector)
        partes.append(f"via {nome}")

    if periodo:
        lbl = PERIODO_LABEL.get(periodo)
        if lbl:
            partes.append(lbl)

    return " ".join(partes)


# ── Filtros (subset dos da listagem de verbatins) ─────────────────────


def _resolver_periodo(periodo: str) -> Optional[datetime]:
    """``7d``/``30d``/``90d``/``6m``/``12m``/``15m`` → datetime início.

    Inválido → None (e o caller devolve 400). Vazio = sem filtro (tudo).
    """
    if not periodo:
        return None
    hoje = datetime.utcnow()
    mapa = {
        "7d": timedelta(days=7),
        "30d": timedelta(days=30),
        "90d": timedelta(days=90),
        "6m": timedelta(days=180),
        "12m": timedelta(days=365),
        "15m": timedelta(days=450),
    }
    delta = mapa.get(periodo)
    if delta is None:
        return None
    return hoje - delta


def _aplicar_filtros(q, empresa_id: int, s):
    """Aplica os filtros do painel na query base.

    Devolve ``(q, erro_response)``. ``erro_response`` é tupla (json, status)
    se houve erro de parsing — caller retorna direto.
    """
    ag_id_raw = request.args.get("agrupamento_id")
    if ag_id_raw:
        try:
            ag_id = int(ag_id_raw)
        except ValueError:
            return q, (jsonify({"erro": "agrupamento_id deve ser inteiro"}), 400)
        locais_do_ag = [
            lid
            for (lid,) in s.query(Local.id)
            .filter_by(empresa_id=empresa_id, agrupamento_id=ag_id)
            .all()
        ]
        if locais_do_ag:
            q = q.filter(Verbatim.local_id.in_(locais_do_ag))
        else:
            q = q.filter(Verbatim.id.is_(None))  # zera o resultado

    local_id_raw = request.args.get("local_id")
    if local_id_raw:
        try:
            q = q.filter(Verbatim.local_id == int(local_id_raw))
        except ValueError:
            return q, (jsonify({"erro": "local_id deve ser inteiro"}), 400)

    fonte_id_raw = request.args.get("fonte_id")
    if fonte_id_raw:
        try:
            q = q.filter(Verbatim.fonte_id == int(fonte_id_raw))
        except ValueError:
            return q, (jsonify({"erro": "fonte_id deve ser inteiro"}), 400)

    periodo = request.args.get("periodo")
    if periodo:
        d = _resolver_periodo(periodo)
        if d is None:
            return q, (
                jsonify({"erro": "periodo inválido. Use: 7d, 30d, 90d, 6m, 12m, 15m"}),
                400,
            )
        q = q.filter(Verbatim.data_criacao_original >= d)

    return q, None


def _filtros_efetivos() -> Dict[str, Any]:
    """Retorna dict serializável dos filtros usados (eco para o front)."""
    return {
        k: request.args.get(k)
        for k in ("agrupamento_id", "local_id", "fonte_id", "periodo")
        if request.args.get(k)
    }


# ── Endpoint Nível 1: 4 pilares ────────────────────────────────────────


@cliente_pode_ver_empresa("empresa_id")
def painel_nivel1(empresa_id: int):
    """Totais por pilar (P, D, Pa, A) + métricas consolidadas (Cap. 4)."""
    with db_session() as s:
        q = s.query(
            Verbatim.subpilar,
            Verbatim.tipo,
            func.count(Verbatim.id),
        ).filter(Verbatim.empresa_id == empresa_id)
        q, erro = _aplicar_filtros(q, empresa_id, s)
        if erro is not None:
            return erro
        q = q.group_by(Verbatim.subpilar, Verbatim.tipo)
        rows = q.all()

        # Agrega por pilar e também monta a matriz por subpilar
        # (mesma estrutura do nivel2) para alimentar Índice/Previsibilidade.
        pilares_agg: Dict[str, Dict[str, int]] = {
            p: {"total": 0, "promotor": 0, "conversivel": 0, "detrator": 0, "inativo": 0}
            for p in PILARES_ORDEM
        }
        subpilares_agg: Dict[str, Dict[str, int]] = {
            sp: {"promotor": 0, "conversivel": 0, "detrator": 0, "inativo": 0, "total": 0}
            for sp in SUBPILARES_ORDEM
        }
        outros = {"sem_lastro": 0, "sem_classificacao": 0}
        total_geral = 0

        for subpilar, tipo, qtd in rows:
            total_geral += qtd
            if subpilar in PILAR_DE_SUBPILAR:
                pilar = PILAR_DE_SUBPILAR[subpilar]
                pilares_agg[pilar]["total"] += qtd
                if tipo in pilares_agg[pilar]:
                    pilares_agg[pilar][tipo] += qtd
                subpilares_agg[subpilar]["total"] += qtd
                if tipo in subpilares_agg[subpilar]:
                    subpilares_agg[subpilar][tipo] += qtd
            elif subpilar == "sem_lastro":
                outros["sem_lastro"] += qtd
            else:
                outros["sem_classificacao"] += qtd

        # Constrói matriz com ratios para alimentar as 3 métricas.
        matriz_para_metricas: List[Dict[str, Any]] = []
        for sp in SUBPILARES_ORDEM:
            cell = subpilares_agg[sp]
            ratio = calcular_ratio(cell["promotor"], cell["detrator"])
            matriz_para_metricas.append({**cell, "subpilar": sp, "ratio": ratio})

        # Pilares construídos antes pra passar ao calcular_indice_geral (opção B).
        pilares: List[Dict[str, Any]] = []
        for p in PILARES_ORDEM:
            agg = pilares_agg[p]
            ratio = calcular_ratio(agg["promotor"], agg["detrator"])
            pilares.append(
                {
                    "pilar": p,
                    "nome": NOME_PILAR[p],
                    "total": agg["total"],
                    "promotor": agg["promotor"],
                    "conversivel": agg["conversivel"],
                    "detrator": agg["detrator"],
                    "inativo": agg["inativo"],
                    "ratio": ratio,
                    "faixa": faixa_ratio(ratio),
                }
            )

        # Resolve data_inicio_periodo para reusar em previsibilidade/concentração
        periodo_arg = request.args.get("periodo")
        data_inicio_periodo = _resolver_periodo(periodo_arg) if periodo_arg else None
        filtros_query = {
            "agrupamento_id": request.args.get("agrupamento_id"),
            "local_id": request.args.get("local_id"),
            "fonte_id": request.args.get("fonte_id"),
            "data_inicio_periodo": data_inicio_periodo,
        }

        # Previsibilidade = dispersão pura (lojas + tempo); None quando sem eixo medível.
        previsibilidade = calcular_previsibilidade(empresa_id, s, filtros_query)
        previsib_medida = previsibilidade_medida(empresa_id, s, filtros_query)  # guard T1
        concentracao_pct = calcular_concentracao_detratores(empresa_id, s, filtros_query)
        conc_n_lojas = concentracao_n_lojas(empresa_id, s, filtros_query)  # guard T2
        # Engajamento (CP-E1): 4º indicador — pré-condição operacional (volume/
        # diversidade/consistência) + selo de confiança por volume.
        engaj = engajamento_escopo(empresa_id, s, filtros_query)

        # Texto descritivo do escopo (hotfix UI 2026-05-24).
        from src.models.agrupamento import Agrupamento
        from src.models.empresa import Empresa as _Empresa
        from src.models.fonte import Fonte as _Fonte

        empresa_obj = s.get(_Empresa, empresa_id)
        empresa_nome = empresa_obj.nome if empresa_obj else f"empresa #{empresa_id}"
        ag_nome = None
        if request.args.get("agrupamento_id"):
            try:
                ag = s.get(Agrupamento, int(request.args.get("agrupamento_id")))
                ag_nome = ag.nome if ag else None
            except (ValueError, TypeError):
                pass
        local_nome = None
        if request.args.get("local_id"):
            try:
                loc = s.get(Local, int(request.args.get("local_id")))
                local_nome = loc.nome if loc else None
            except (ValueError, TypeError):
                pass
        fonte_conector = None
        if request.args.get("fonte_id"):
            try:
                fonte = s.get(_Fonte, int(request.args.get("fonte_id")))
                fonte_conector = fonte.conector_tipo if fonte else None
            except (ValueError, TypeError):
                pass

        filtros_descricao = descrever_escopo(
            empresa_nome,
            agrupamento_nome=ag_nome,
            local_nome=local_nome,
            fonte_conector=fonte_conector,
            periodo=periodo_arg,
        )

    # Índice Geral (Manual v8: min(pior_pilar, média) → normalização por partes).
    indice_geral = calcular_indice_geral(matriz_para_metricas, pilares=pilares)
    indice_governado = indice_governado_pelo_pior(matriz_para_metricas, pilares=pilares)

    # Índice PDPA (Manual Cap. 4): a relação em um número + Base (P+D) / Topo (Pa+A).
    pdpa_geral, _ = indice_pdpa(pilares)
    pdpa_base, pdpa_base_vol = indice_pdpa(pilares, codigos={"P", "D"})
    pdpa_topo, pdpa_topo_vol = indice_pdpa(pilares, codigos={"Pa", "A"})

    return jsonify(
        {
            "empresa_id": empresa_id,
            "filtros": _filtros_efetivos(),
            "total_verbatins": total_geral,
            "pilares": pilares,
            "outros": outros,
            # B5 ext. CP-3: métricas consolidadas (Manual Cap. 4)
            "indice_geral": indice_geral,
            "indice_geral_faixa": faixa_indice_geral(indice_geral),
            "indice_geral_governado_pelo_pior": indice_governado,  # nota do card (pior binda)
            # Índice PDPA (manchete): a relação em um número + Base/Topo com volume.
            "indice_pdpa": pdpa_geral,
            "indice_pdpa_base": pdpa_base,
            "indice_pdpa_base_volume": pdpa_base_vol,
            "indice_pdpa_topo": pdpa_topo,
            "indice_pdpa_topo_volume": pdpa_topo_vol,
            "previsibilidade": previsibilidade,
            "previsibilidade_medida": previsib_medida,  # guard T1 (default 70,0 vs medido)
            "concentracao_detratores": concentracao_pct,
            "concentracao_faixa": faixa_concentracao(concentracao_pct),
            "concentracao_n_lojas": conc_n_lojas,  # guard T2 (top-5 trivial c/ poucas lojas)
            # Engajamento (CP-E1): índice 0-100 + componentes + selo de confiança.
            "indice_engajamento": engaj["indice"],
            "engajamento_componentes": engaj["componentes"],
            "engajamento_selo": engaj["selo"],
            "engajamento_selo_emoji": engaj["selo_emoji"],
            "engajamento_volume": engaj["volume"],
            "engajamento_fontes_ativas": engaj["fontes_ativas"],
            "engajamento_fontes_cadastradas": engaj["fontes_cadastradas"],
            # Hotfix UI: texto descritivo do escopo p/ os 3 cards
            "filtros_descricao": filtros_descricao,
        }
    )


# ── Endpoint Nível 2: matriz subpilar × tipo ──────────────────────────


@cliente_pode_ver_empresa("empresa_id")
def painel_nivel2(empresa_id: int):
    """Matriz 12 subpilares × 3 tipos (promotor/conversivel/detrator).

    ``inativo`` aparece como coluna informativa porque ``sem_lastro`` vai
    junto, mas a matriz principal são os 12 subpilares P/D/Pa/A.
    """
    with db_session() as s:
        q = s.query(
            Verbatim.subpilar,
            Verbatim.tipo,
            func.count(Verbatim.id),
        ).filter(Verbatim.empresa_id == empresa_id)
        q, erro = _aplicar_filtros(q, empresa_id, s)
        if erro is not None:
            return erro
        q = q.group_by(Verbatim.subpilar, Verbatim.tipo)
        rows = q.all()

    matriz_agg: Dict[str, Dict[str, int]] = {
        sp: {"promotor": 0, "conversivel": 0, "detrator": 0, "inativo": 0, "total": 0}
        for sp in SUBPILARES_ORDEM
    }
    sem_lastro_agg = {"promotor": 0, "conversivel": 0, "detrator": 0, "inativo": 0, "total": 0}
    sem_classif_agg = {"promotor": 0, "conversivel": 0, "detrator": 0, "inativo": 0, "total": 0}
    total_geral = 0

    for subpilar, tipo, qtd in rows:
        total_geral += qtd
        if subpilar in matriz_agg:
            matriz_agg[subpilar]["total"] += qtd
            if tipo in matriz_agg[subpilar]:
                matriz_agg[subpilar][tipo] += qtd
        elif subpilar == "sem_lastro":
            sem_lastro_agg["total"] += qtd
            if tipo in sem_lastro_agg:
                sem_lastro_agg[tipo] += qtd
        else:
            sem_classif_agg["total"] += qtd
            if tipo in sem_classif_agg:
                sem_classif_agg[tipo] += qtd

    matriz: List[Dict[str, Any]] = []
    for sp in SUBPILARES_ORDEM:
        cell = matriz_agg[sp]
        ratio = calcular_ratio(cell["promotor"], cell["detrator"])
        matriz.append(
            {
                "subpilar": sp,
                "nome": NOME_SUBPILAR[sp],
                "pilar": PILAR_DE_SUBPILAR[sp],
                "promotor": cell["promotor"],
                "conversivel": cell["conversivel"],
                "detrator": cell["detrator"],
                "inativo": cell["inativo"],
                "total": cell["total"],
                "ratio": ratio,
                "faixa": faixa_ratio(ratio),
            }
        )

    return jsonify(
        {
            "empresa_id": empresa_id,
            "filtros": _filtros_efetivos(),
            "total_verbatins": total_geral,
            "matriz": matriz,
            "sem_lastro": sem_lastro_agg,
            "sem_classificacao": sem_classif_agg,
        }
    )


# ── Leitura textual sequencial (Bloco 5 ext. CP-5) ────────────────────


@cliente_pode_ver_empresa("empresa_id")
def painel_leitura(empresa_id: int):
    """Leitura textual sequencial via Sonnet (Manual Cap. 3).

    Chama painel_nivel1 internamente para obter o estado atual, depois
    pede ao Sonnet para interpretar em 2-3 frases. Carregado async no
    UI para não atrasar o painel principal.
    """
    from src.api.painel_leitura import gerar_leitura_sequencial

    resp_n1 = painel_nivel1(empresa_id)
    if isinstance(resp_n1, tuple):
        return resp_n1  # propaga erro (400/403)
    n1 = resp_n1.get_json()
    leitura = gerar_leitura_sequencial(n1 or {})
    return jsonify({"empresa_id": empresa_id, "leitura": leitura})


# ── Exportar XLSX (Bloco 5 CP-3) ──────────────────────────────────────


@cliente_pode_ver_empresa("empresa_id")
def exportar_painel_xlsx(empresa_id: int):
    """Exporta painel (Visão Geral + Detalhamento) em XLSX com 2 sheets."""
    from io import BytesIO

    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # Reusa a lógica dos 2 endpoints chamando-os internamente
    resp_n1 = painel_nivel1(empresa_id)
    if isinstance(resp_n1, tuple):
        return resp_n1
    resp_n2 = painel_nivel2(empresa_id)
    if isinstance(resp_n2, tuple):
        return resp_n2
    n1 = resp_n1.get_json()
    n2 = resp_n2.get_json()

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")

    # Sheet 1: Visão Geral
    ws1 = wb.active
    ws1.title = "Visão Geral"
    ws1.append([f"Empresa #{empresa_id} — Painel Executivo (Visão Geral)"])
    ws1["A1"].font = bold
    filtros = n1.get("filtros") or {}
    if filtros:
        ws1.append(["Filtros aplicados:", " | ".join(f"{k}={v}" for k, v in filtros.items())])
    ws1.append([f"Total verbatins: {n1.get('total_verbatins', 0)}"])
    ws1.append([])
    headers1 = [
        "Pilar",
        "Nome",
        "Total",
        "Promotor",
        "Conversível",
        "Detrator",
        "Inativo",
        "Ratio P/D",
        "Faixa",
    ]
    ws1.append(headers1)
    for cell in ws1[ws1.max_row]:
        cell.font = bold
        cell.fill = header_fill
    for p in n1.get("pilares", []):
        ws1.append(
            [
                p["pilar"],
                p["nome"],
                p["total"],
                p["promotor"],
                p["conversivel"],
                p["detrator"],
                p["inativo"],
                p.get("ratio", 0.0),
                p.get("faixa", ""),
            ]
        )
    ws1.append([])
    outros = n1.get("outros") or {}
    if outros:
        ws1.append(["Fora dos 4 pilares:"])
        ws1[ws1.max_row][0].font = bold
        ws1.append(["sem_lastro", outros.get("sem_lastro", 0)])
        ws1.append(["sem_classificação", outros.get("sem_classificacao", 0)])

    # Sheet 2: Detalhamento por Subpilar
    ws2 = wb.create_sheet("Detalhamento por Subpilar")
    ws2.append([f"Empresa #{empresa_id} — Detalhamento por Subpilar"])
    ws2["A1"].font = bold
    if filtros:
        ws2.append(["Filtros aplicados:", " | ".join(f"{k}={v}" for k, v in filtros.items())])
    ws2.append([])
    headers2 = [
        "Pilar",
        "Subpilar",
        "Nome do Subpilar",
        "Promotor",
        "Conversível",
        "Detrator",
        "Inativo",
        "Total",
        "Ratio P/D",
        "Faixa",
    ]
    ws2.append(headers2)
    for cell in ws2[ws2.max_row]:
        cell.font = bold
        cell.fill = header_fill
    for c in n2.get("matriz", []):
        ws2.append(
            [
                c["pilar"],
                c["subpilar"],
                c.get("nome", ""),
                c["promotor"],
                c["conversivel"],
                c["detrator"],
                c["inativo"],
                c["total"],
                c.get("ratio", 0.0),
                c.get("faixa", ""),
            ]
        )
    sl = n2.get("sem_lastro") or {}
    sc = n2.get("sem_classificacao") or {}
    if sl.get("total"):
        ws2.append(
            [
                "—",
                "sem_lastro",
                "(sem ancoragem)",
                "—",
                "—",
                "—",
                sl.get("inativo", 0),
                sl["total"],
                "—",
                "—",
            ]
        )
    if sc.get("total"):
        ws2.append(
            [
                "—",
                "sem classificação",
                "(falha classifier)",
                "—",
                "—",
                "—",
                "—",
                sc["total"],
                "—",
                "—",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"painel_empresa_{empresa_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
